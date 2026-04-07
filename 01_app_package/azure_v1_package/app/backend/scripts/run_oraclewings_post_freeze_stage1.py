#!/usr/bin/env python3
import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None

BASE_DIR = Path("/Users/integrationwings/Desktop/LLM_Wrap/iwerp-prod")
WORKSPACE_DIR = BASE_DIR.parent
ORACLEWINGS_DIR = WORKSPACE_DIR / "Oraclewings_ai"
OUTPUT_DIR = BASE_DIR / "specialization_tracks" / "oraclewings_post_freeze_stage1"
MANIFEST_DIR = OUTPUT_DIR / "manifests"
INDEXES_DIR = BASE_DIR / "backend" / "core" / "retrieval" / "vectors"

sys.path.append(str(BASE_DIR))
sys.path.append(str(BASE_DIR / "backend"))

from backend.core.ingest.curation import CuratedIngestionValidator, stable_hash  # noqa: E402
from backend.core.retrieval.vectors.faiss_index import FaissIndex  # noqa: E402
from backend.core.schemas.curation import (  # noqa: E402
    CorpusType,
    DocType,
    IngestionManifestRecord,
    SourceSystem,
)


FASTFORMULA_KT_PATH = ORACLEWINGS_DIR / "backend" / "backend_hcm" / "Fastformula_KT.csv"
FORMULA_TYPES_PATH = ORACLEWINGS_DIR / "backend" / "backend_hcm" / "app" / "data" / "formula_types.json"
DBI_LIBRARY_PATHS = [
    ORACLEWINGS_DIR / "backend" / "backend_hcm" / "app" / "data" / "dbi_library_part1.json",
    ORACLEWINGS_DIR / "backend" / "backend_hcm" / "app" / "data" / "dbi_library_part2.json",
    ORACLEWINGS_DIR / "backend" / "backend_hcm" / "app" / "data" / "dbi_library_part3.json",
]

MODULES_TABLES_PATH = ORACLEWINGS_DIR / "backend" / "orawing_ai" / "data" / "metadata" / "modules_tables.json"
JOIN_HINTS_PATH = ORACLEWINGS_DIR / "backend" / "orawing_ai" / "data" / "metadata" / "join_hints.json"
COLUMN_SEMANTICS_PATH = ORACLEWINGS_DIR / "backend" / "orawing_ai" / "data" / "metadata" / "column_semantics.json"
VIEW_EXPANSIONS_PATH = ORACLEWINGS_DIR / "backend" / "orawing_ai" / "data" / "metadata" / "view_expansions.json"
DDL_FILES = [
    ORACLEWINGS_DIR / "database" / "schema" / "financials_schema.sql",
    ORACLEWINGS_DIR / "database" / "schema" / "hcm_schema.sql",
    ORACLEWINGS_DIR / "database" / "schema" / "procurement_schema.sql",
    ORACLEWINGS_DIR / "database" / "schema" / "scm_schema.sql",
    ORACLEWINGS_DIR / "database" / "schema" / "project_management_schema.sql",
]
CRAWLER_METADATA_DIRS = [
    ORACLEWINGS_DIR / "backend" / "orawing_ai" / "data" / "crawler" / "metadata" / "metadata",
    ORACLEWINGS_DIR / "backend" / "orawing_ai" / "data" / "crawler" / "metadata",
]
CRAWLER_METADATA_CSV_FILES = [
    "all_table_RPT.csv",
    "all_table_rpt2.csv",
    "all_table_rpt3.csv",
    "all_tab_columns_rpt.csv",
    "ALL_TABLES.csv",
    "ALL_TAB_COLUMNS.csv",
    "ALL_VIEWS_RPT.csv",
    "all_views .csv",
    "all_views  (1).csv",
    "FUSION_FK_GRAPH.csv",
    "ALL_CONS_COLUMNS_AND_TABLE_RPT.csv",
    "all_cons_columns_Extract.csv",
    "all_ind_columns.csv",
    "all_indexes_Extract.csv",
    "ALL_SYNONYMS.csv",
    "all_table_dm4_test.csv",
    "all_table_dm5.csv",
    "all_table_Rpt1.csv",
]
CRAWLER_METADATA_XLSX_FILES = [
    "all_constraints_Extract.xlsx",
    "ALL_CONSTRAINTS_FULL_ALL_CONSTRAINTS_FULL.xlsx",
]
FULL_SQL_SCAN_TEXT_EXTENSIONS = {
    ".sql",
    ".json",
    ".md",
    ".txt",
    ".py",
    ".yaml",
    ".yml",
    ".csv",
    ".xml",
}
FULL_SQL_SCAN_EXCLUDED_DIRS = {
    ".git",
    "__pycache__",
    "node_modules",
    ".venv",
    "venv",
    ".pytest_cache",
    ".mypy_cache",
    ".idea",
    ".vscode",
}
FULL_SQL_SCAN_MAX_FILE_BYTES = 4_000_000
FULL_SQL_SCAN_MAX_ACCEPTED_RECORDS = 500

TROUBLESHOOTING_KB_PATH = ORACLEWINGS_DIR / "backend" / "scm_bot_backend" / "knowledge_base" / "troubleshooting_kb.json"

EXCLUDED_SOURCE_HINTS = (".db", ".sqlite", ".pyc", ".pack", ".idx", "tmp_snip", "test_crawl_output")


@dataclass
class BuildContext:
    generated_at_utc: str
    tenant: str
    index_enabled: bool


def _normalize(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def _slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


def _read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8", errors="ignore"))


def _read_lines_from_single_column_csv(path: Path) -> List[str]:
    lines: List[str] = []
    with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
        reader = csv.reader(handle)
        next(reader, None)
        for row in reader:
            if row and row[0].strip():
                lines.append(row[0].strip())
    return lines


def _is_git_lfs_pointer(path: Path) -> bool:
    try:
        with path.open("r", encoding="utf-8", errors="ignore") as handle:
            first_line = (handle.readline() or "").strip().lower()
        return first_line.startswith("version https://git-lfs.github.com/spec/v1")
    except OSError:
        return False


def _normalize_header(value: str) -> str:
    cleaned = re.sub(r"[^A-Z0-9]+", "_", (value or "").strip().lstrip("\ufeff").upper()).strip("_")
    return cleaned


def _normalize_headers(headers: List[str]) -> List[str]:
    normalized: List[str] = []
    seen: Counter[str] = Counter()
    for header in headers:
        base = _normalize_header(str(header))
        if not base:
            normalized.append("")
            continue
        seen[base] += 1
        normalized.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return normalized


def _clean_schema_identifier(value: Any) -> str:
    text = str(value or "").strip().strip('"').upper()
    if not text:
        return ""
    if text in {"NA", "N/A", "NULL", "NONE", "-"}:
        return ""
    token = text.split(".")[-1].strip()
    if not re.match(r"^[A-Z][A-Z0-9_$#]*$", token):
        return ""
    return token


def _iter_csv_rows(path: Path) -> Iterable[Dict[str, str]]:
    with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
        first_line = handle.readline()
        if not first_line:
            return
        if first_line.strip().lower().startswith("version https://git-lfs.github.com/spec/v1"):
            raise ValueError("git_lfs_pointer")
        delimiter = "|" if first_line.count("|") > first_line.count(",") else ","
        handle.seek(0)
        reader = csv.reader(handle, delimiter=delimiter)
        raw_headers = next(reader, [])
        headers = _normalize_headers([str(h) for h in raw_headers])
        if not any(headers):
            return
        for raw_row in reader:
            if not raw_row or not any(str(cell).strip() for cell in raw_row):
                continue
            row = list(raw_row)
            if len(row) < len(headers):
                row.extend([""] * (len(headers) - len(row)))
            payload: Dict[str, str] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                payload[header] = str(row[idx]).strip()
            if payload:
                yield payload


def _iter_xlsx_rows(path: Path) -> Iterable[Dict[str, str]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl_not_available")
    wb = load_workbook(path, read_only=True, data_only=True)
    for ws in wb.worksheets:
        headers: Optional[List[str]] = None
        header_row_idx = 0
        for row_idx, raw_row in enumerate(ws.iter_rows(values_only=True), start=1):
            row = ["" if val is None else str(val).strip() for val in raw_row]
            if headers is None:
                candidate = _normalize_headers(row)
                if {"TABLE_NAME", "CONSTRAINT_NAME"} & set(candidate):
                    headers = candidate
                    header_row_idx = row_idx
                    continue
                if row_idx >= 40:
                    break
                continue
            if row_idx <= header_row_idx:
                continue
            if not any(row):
                continue
            payload: Dict[str, str] = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = row[idx] if idx < len(row) else ""
                payload[header] = value
            if payload:
                yield payload


def _extract_base_tables_from_view_sql(text: str) -> List[str]:
    tables: List[str] = []
    if not text:
        return tables
    for token in re.findall(r'(?is)\b(?:FROM|JOIN)\s+"?([A-Za-z0-9_$#]+)"?', text):
        cleaned = _clean_schema_identifier(token)
        if cleaned:
            tables.append(cleaned)
    return tables


def _relative_or_abs(path: Path) -> str:
    try:
        return str(path.relative_to(ORACLEWINGS_DIR))
    except ValueError:
        return str(path)


def _discover_crawler_metadata_sources() -> List[Path]:
    seen: set[str] = set()
    discovered: List[Path] = []
    for directory in CRAWLER_METADATA_DIRS:
        if not directory.exists():
            continue
        for name in CRAWLER_METADATA_CSV_FILES + CRAWLER_METADATA_XLSX_FILES:
            path = directory / name
            if not path.exists():
                continue
            key = str(path.resolve())
            if key in seen:
                continue
            seen.add(key)
            discovered.append(path)
    return discovered


def _iter_full_repo_scan_files(root: Path) -> Iterable[Path]:
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        parts = set(path.parts)
        if FULL_SQL_SCAN_EXCLUDED_DIRS & parts:
            continue
        if path.suffix.lower() not in FULL_SQL_SCAN_TEXT_EXTENSIONS:
            continue
        yield path


def _is_probably_unsupported_sql(sql: str) -> Optional[str]:
    normalized = sql.strip()
    if not normalized:
        return "empty_sql"
    if re.search(r"(?i)\b(prompt|spool|set\s+\w+|exit)\b", normalized):
        return "sqlplus_control_sql"
    if re.search(r":[A-Za-z][A-Za-z0-9_]*", normalized):
        return "parameterized_placeholder_sql"
    if re.search(r"\$\{[^}]+\}|<[^>\n]{1,80}>", normalized):
        return "template_placeholder_sql"
    if len(normalized) < 40:
        return "sql_too_short"
    return CuratedIngestionValidator.reject_sql(normalized)


def _split_candidate_sql_blocks(text: str) -> List[str]:
    blocks: List[str] = []

    for m in re.finditer(r"(?is)```(?:sql)?\s*(.+?)```", text):
        candidate = m.group(1).strip()
        if candidate:
            blocks.append(candidate)

    statement_pattern = re.compile(
        r"(?is)\bselect\b[\s\S]{0,5000}?\bfrom\b[\s\S]{0,5000}?(?:;|$)"
    )
    for m in statement_pattern.finditer(text):
        candidate = m.group(0).strip()
        if candidate:
            blocks.append(candidate)

    unique: List[str] = []
    seen: set[str] = set()
    for block in blocks:
        normalized = _normalize(block)
        if normalized in seen:
            continue
        seen.add(normalized)
        unique.append(block.strip())
    return unique


def _extract_tables_from_sql(sql: str) -> List[str]:
    tables: List[str] = []
    for token in re.findall(r'(?is)\b(?:FROM|JOIN)\s+"?([A-Za-z0-9_$#\.]+)"?', sql):
        cleaned = _clean_schema_identifier(token)
        if cleaned:
            tables.append(cleaned)
    deduped: List[str] = []
    seen: set[str] = set()
    for name in tables:
        if name in seen:
            continue
        seen.add(name)
        deduped.append(name)
    return deduped


def _infer_module_from_tables(tables: List[str]) -> str:
    if not tables:
        return "Financials"
    counts: Counter[str] = Counter(_infer_module_from_table_name(name) for name in tables)
    return counts.most_common(1)[0][0]


def _load_allowed_oracle_sql_objects() -> set[str]:
    allowed: set[str] = set()

    try:
        modules_tables = _read_json(MODULES_TABLES_PATH)
        if isinstance(modules_tables, dict):
            for values in modules_tables.values():
                if not isinstance(values, list):
                    continue
                for value in values:
                    token = _clean_schema_identifier(value)
                    if token:
                        allowed.add(token)
    except Exception:
        pass

    crawler_top = ORACLEWINGS_DIR / "backend" / "orawing_ai" / "data" / "crawler" / "metadata"
    for filename, key in (("ALL_TABLES.csv", "TABLE_NAME"), ("ALL_VIEWS_RPT.csv", "VIEW_NAME")):
        path = crawler_top / filename
        if not path.exists() or _is_git_lfs_pointer(path):
            continue
        try:
            for row in _iter_csv_rows(path):
                token = _clean_schema_identifier(row.get(key))
                if token:
                    allowed.add(token)
        except Exception:
            continue

    for ddl_path in DDL_FILES:
        if not ddl_path.exists():
            continue
        try:
            tables, views = _parse_ddl_objects(ddl_path)
        except Exception:
            continue
        for name in tables + views:
            token = _clean_schema_identifier(name)
            if token:
                allowed.add(token)

    return allowed


def _oraclewings_uri(rel_path: str, anchor: Optional[str] = None) -> str:
    if anchor:
        return f"oraclewings://{rel_path}#{anchor}"
    return f"oraclewings://{rel_path}"


def _extract_formula_dbis(content: str) -> List[str]:
    candidates = set(re.findall(r"\b([A-Z]{2,}_[A-Z0-9_]{2,})\b", content))
    reserved = {"DEFAULT", "INPUTS", "RETURN", "THEN", "ELSE", "END", "IF", "DATE", "CALC"}
    return sorted(c for c in candidates if c.split("_")[0] not in reserved)


def _infer_formula_type(title: str, content: str) -> str:
    text = f"{title} {content}".lower()
    mapping = [
        ("proration", "Proration"),
        ("accrual", "Accrual"),
        ("absence", "Absence"),
        ("validation", "Validation"),
        ("rate", "Rate"),
        ("payroll", "Payroll"),
        ("extract", "Oracle HCM Extract Rules"),
        ("eligibility", "Eligibility"),
    ]
    for token, label in mapping:
        if token in text:
            return label
    return "General"


def _manifest_record(
    *,
    source_path: str,
    title: str,
    module: str,
    task_type: str,
    doc_type: DocType,
    corpus: CorpusType,
    quality_score: float,
    content: str,
    trusted_schema_objects: Optional[List[str]] = None,
    metadata: Optional[Dict[str, Any]] = None,
    source_uri: Optional[str] = None,
) -> Dict[str, Any]:
    content_norm = _normalize(content)
    md = dict(metadata or {})
    md.setdefault("source_uri", source_uri or source_path)
    md.setdefault("stage", "oraclewings_post_freeze_stage1")
    md.setdefault("title", title)
    md.setdefault("module", module)
    md.setdefault("task_type", task_type)
    md.setdefault("doc_type", doc_type.value)
    md.setdefault("quality_score", quality_score)
    trusted = trusted_schema_objects or []
    return {
        "source_path": source_path,
        "source_uri": source_uri or source_path,
        "title": title,
        "module": module,
        "task_type": task_type,
        "doc_type": doc_type.value,
        "trusted_schema_objects": trusted,
        "quality_score": quality_score,
        "content_hash": stable_hash(corpus.value, source_path, title, content_norm),
        "source_system": SourceSystem.ORACLEWINGS_REPO.value,
        "corpus": corpus.value,
        "content": content_norm,
        "metadata": md,
    }


def _validate_record(payload: Dict[str, Any]) -> Tuple[bool, Optional[str]]:
    record = IngestionManifestRecord(**payload)
    return CuratedIngestionValidator.validate_manifest_record(record)


def _fast_formula_stage1() -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    rejects: List[Dict[str, Any]] = []

    lines = _read_lines_from_single_column_csv(FASTFORMULA_KT_PATH)

    def append_formula_block(title: str, block_lines: List[str], section_id: str, is_example: bool) -> None:
        content = "\n".join(block_lines).strip()
        if not content:
            return
        formula_type = _infer_formula_type(title, content)
        dbis = _extract_formula_dbis(content)
        task_type = "fast_formula_generation"
        quality = 0.92 if is_example else 0.84
        rec = _manifest_record(
            source_path=str(FASTFORMULA_KT_PATH.relative_to(ORACLEWINGS_DIR)),
            source_uri=_oraclewings_uri(
                str(FASTFORMULA_KT_PATH.relative_to(ORACLEWINGS_DIR)),
                anchor=section_id,
            ),
            title=title,
            module="HCM",
            task_type=task_type,
            doc_type=DocType.FAST_FORMULA_EXAMPLE,
            corpus=CorpusType.FAST_FORMULA,
            quality_score=quality,
            content=content,
            trusted_schema_objects=[],
            metadata={
                "formula_type": formula_type,
                "dbis": dbis,
                "source_section": section_id,
                "is_template": not is_example,
                "is_full_example": is_example,
            },
        )
        ok, reason = _validate_record(rec)
        if ok:
            records.append(rec)
        else:
            rejects.append(
                {
                    "source_path": rec["source_path"],
                    "title": title,
                    "reason": reason or "validation_failed",
                    "asset_type": "fast_formula",
                }
            )

    # 4.x templates
    for idx, line in enumerate(lines):
        m = re.match(r"^4\.(\d+)\s+Template\s+[—-]\s+(.+)$", line, flags=re.IGNORECASE)
        if not m:
            continue
        num = m.group(1)
        title = f"Template {num}: {m.group(2)}"
        block: List[str] = [line]
        j = idx + 1
        while j < len(lines):
            nxt = lines[j]
            if re.match(r"^4\.\d+\s+Template", nxt, flags=re.IGNORECASE):
                break
            if nxt.startswith("5. FULL FORMULA EXAMPLES"):
                break
            block.append(nxt)
            j += 1
        append_formula_block(title, block, f"template_{num}", is_example=False)

    # full examples in section 5
    example_idxs = [
        (i, re.match(r"^✅\s*Example\s*(\d+)\s+[—-]\s+(.+)$", l))
        for i, l in enumerate(lines)
        if re.match(r"^✅\s*Example\s*(\d+)\s+[—-]\s+(.+)$", l)
    ]
    for pos, (idx, match) in enumerate(example_idxs):
        if not match:
            continue
        ex_num = match.group(1)
        title = f"Full Example {ex_num}: {match.group(2)}"
        start = idx
        end = example_idxs[pos + 1][0] if pos + 1 < len(example_idxs) else len(lines)
        block = lines[start:end]
        append_formula_block(title, block, f"example_{ex_num}", is_example=True)

    # formula types catalog
    formula_types = _read_json(FORMULA_TYPES_PATH)
    if isinstance(formula_types, list) and formula_types:
        rec = _manifest_record(
            source_path=str(FORMULA_TYPES_PATH.relative_to(ORACLEWINGS_DIR)),
            source_uri=_oraclewings_uri(str(FORMULA_TYPES_PATH.relative_to(ORACLEWINGS_DIR))),
            title="Fast Formula Type Catalog",
            module="HCM",
            task_type="fast_formula_generation",
            doc_type=DocType.FAST_FORMULA_DOC,
            corpus=CorpusType.FAST_FORMULA,
            quality_score=0.9,
            content="\n".join(f"- {t}" for t in formula_types),
            trusted_schema_objects=[],
            metadata={
                "formula_type_count": len(formula_types),
                "formula_types": formula_types,
                "catalog_kind": "formula_types",
            },
        )
        ok, reason = _validate_record(rec)
        if ok:
            records.append(rec)
        else:
            rejects.append(
                {
                    "source_path": rec["source_path"],
                    "title": rec["title"],
                    "reason": reason or "validation_failed",
                    "asset_type": "fast_formula",
                }
            )

    # DBI library aggregation (safe summary, not raw 466k row ingest)
    prefix_counter: Counter[str] = Counter()
    dtype_counter: Counter[str] = Counter()
    prefix_samples: Dict[str, List[str]] = defaultdict(list)
    total_dbi_rows = 0

    for path in DBI_LIBRARY_PATHS:
        rows = _read_json(path)
        if not isinstance(rows, list):
            continue
        for row in rows:
            total_dbi_rows += 1
            name = str(row.get("name", "")).upper().strip()
            dtype = str(row.get("data_type", "")).upper().strip() or "UNKNOWN"
            dtype_counter[dtype] += 1
            m = re.match(r"^([A-Z0-9]+?)(?:_|$)", name)
            prefix = m.group(1) if m else "UNKNOWN"
            prefix_counter[prefix] += 1
            sample_bucket = prefix_samples[prefix]
            if len(sample_bucket) < 20 and name:
                sample_bucket.append(name)

    # one summary record
    top_prefixes = prefix_counter.most_common(80)
    summary_lines = [f"Total DBIs: {total_dbi_rows}", "Top DBI prefixes:"]
    for pref, cnt in top_prefixes[:40]:
        summary_lines.append(f"- {pref}: {cnt}")
    summary_lines.append("Data types:")
    for dtype, cnt in dtype_counter.most_common(12):
        summary_lines.append(f"- {dtype}: {cnt}")

    summary_rec = _manifest_record(
        source_path="backend/backend_hcm/app/data/dbi_library_part*.json",
        source_uri=_oraclewings_uri("backend/backend_hcm/app/data/dbi_library_part*.json"),
        title="Fast Formula DBI Library Summary",
        module="HCM",
        task_type="fast_formula_troubleshooting",
        doc_type=DocType.FAST_FORMULA_DOC,
        corpus=CorpusType.FAST_FORMULA,
        quality_score=0.86,
        content="\n".join(summary_lines),
        trusted_schema_objects=[],
        metadata={
            "total_dbi_rows": total_dbi_rows,
            "distinct_prefixes": len(prefix_counter),
            "top_prefixes": top_prefixes[:80],
            "data_types": dict(dtype_counter.most_common()),
            "ingestion_mode": "aggregated_summary_only",
        },
    )
    ok, reason = _validate_record(summary_rec)
    if ok:
        records.append(summary_rec)
    else:
        rejects.append(
            {
                "source_path": summary_rec["source_path"],
                "title": summary_rec["title"],
                "reason": reason or "validation_failed",
                "asset_type": "fast_formula",
            }
        )

    # top prefix records
    for pref, cnt in top_prefixes[:60]:
        sample_names = prefix_samples.get(pref, [])
        content = "\n".join(
            [
                f"DBI Prefix: {pref}",
                f"Count: {cnt}",
                "Sample DBIs:",
                *[f"- {n}" for n in sample_names],
            ]
        )
        rec = _manifest_record(
            source_path="backend/backend_hcm/app/data/dbi_library_part*.json",
            source_uri=_oraclewings_uri("backend/backend_hcm/app/data/dbi_library_part*.json", anchor=f"prefix_{pref}"),
            title=f"Fast Formula DBI Prefix {pref}",
            module="HCM",
            task_type="fast_formula_troubleshooting",
            doc_type=DocType.FAST_FORMULA_DOC,
            corpus=CorpusType.FAST_FORMULA,
            quality_score=0.82,
            content=content,
            trusted_schema_objects=[],
            metadata={
                "dbi_prefix": pref,
                "dbi_count": cnt,
                "sample_names": sample_names,
                "ingestion_mode": "prefix_summary",
            },
        )
        ok, reason = _validate_record(rec)
        if ok:
            records.append(rec)
        else:
            rejects.append(
                {
                    "source_path": rec["source_path"],
                    "title": rec["title"],
                    "reason": reason or "validation_failed",
                    "asset_type": "fast_formula",
                }
            )

    notes = {
        "sources": [
            str(FASTFORMULA_KT_PATH.relative_to(ORACLEWINGS_DIR)),
            str(FORMULA_TYPES_PATH.relative_to(ORACLEWINGS_DIR)),
            *[str(p.relative_to(ORACLEWINGS_DIR)) for p in DBI_LIBRARY_PATHS],
        ],
        "kt_line_count": len(lines),
        "dbi_total_rows": total_dbi_rows,
        "dbi_prefix_count": len(prefix_counter),
    }
    return records, rejects, notes


def _infer_module_from_modules_key(key: str) -> str:
    lowered = key.lower().replace("_", " ")
    if "cash management" in lowered:
        return "Cash Management"
    if "asset" in lowered:
        return "Assets"
    if "tax" in lowered:
        return "Tax"
    if "receivable" in lowered:
        return "Receivables"
    if "payable" in lowered:
        return "Payables"
    if "general ledger" in lowered:
        return "General Ledger"
    if "expense" in lowered:
        return "Expenses"
    if "procurement" in lowered or "purchase" in lowered:
        return "Procurement"
    if "hcm" in lowered:
        return "HCM"
    if "project" in lowered:
        return "Projects"
    return "Financials"


def _infer_module_from_table_name(name: str) -> str:
    upper = name.upper()
    if upper.startswith(("PER_", "PAY_", "ANC_", "IRC_", "HWM_")):
        return "HCM"
    if upper.startswith(("AP_", "IBY_")):
        return "Payables"
    if upper.startswith(("AR_", "RA_", "HZ_")):
        return "Receivables"
    if upper.startswith(("GL_", "XLA_")):
        return "General Ledger"
    if upper.startswith(("PO_", "POR_", "POZ_")):
        return "Procurement"
    if upper.startswith(("INV_", "DOO_", "WSH_", "MSC_", "CST_")):
        return "SCM"
    if upper.startswith(("FA_",)):
        return "Assets"
    return "Financials"


def _parse_ddl_objects(path: Path) -> Tuple[List[str], List[str]]:
    text = path.read_text(encoding="utf-8", errors="ignore")
    pattern = re.compile(
        r"(?im)^\s*CREATE\s+(TABLE|VIEW)\s+((?:\"?[A-Za-z0-9_]+\"?\.)?\"?[A-Za-z0-9_]+\"?)"
    )
    tables: List[str] = []
    views: List[str] = []
    for kind, raw_name in pattern.findall(text):
        name = raw_name.split(".")[-1].replace('"', "").strip()
        if not name:
            continue
        if kind.upper() == "TABLE":
            tables.append(name)
        else:
            views.append(name)
    return tables, views


def _sql_metadata_crawler_stage1() -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    rejects: List[Dict[str, Any]] = []

    discovered_sources = _discover_crawler_metadata_sources()
    if not discovered_sources:
        return records, rejects, {"sources": [], "note": "no_crawler_metadata_sources_discovered"}

    tables: set[str] = set()
    views: set[str] = set()
    columns_by_object: Dict[str, set[str]] = defaultdict(set)
    view_to_base_tables: Dict[str, set[str]] = defaultdict(set)
    fk_edges: set[Tuple[str, str, str, str, str]] = set()
    source_stats: Dict[str, Dict[str, Any]] = {}

    for path in sorted(discovered_sources, key=lambda p: _relative_or_abs(p)):
        rel_path = _relative_or_abs(path)
        ext = path.suffix.lower()
        row_count = 0
        lfs_pointer = _is_git_lfs_pointer(path)
        if lfs_pointer:
            rejects.append(
                {
                    "source_path": rel_path,
                    "title": path.name,
                    "reason": "git_lfs_pointer_not_ingestable",
                    "asset_type": "sql_metadata",
                }
            )
            source_stats[rel_path] = {"rows_seen": 0, "status": "rejected_lfs_pointer"}
            continue

        try:
            if ext == ".csv":
                row_iter = _iter_csv_rows(path)
            elif ext == ".xlsx":
                row_iter = _iter_xlsx_rows(path)
            else:
                rejects.append(
                    {
                        "source_path": rel_path,
                        "title": path.name,
                        "reason": "unsupported_metadata_extension",
                        "asset_type": "sql_metadata",
                    }
                )
                source_stats[rel_path] = {"rows_seen": 0, "status": "unsupported_extension"}
                continue

            for row in row_iter:
                row_count += 1
                table_name = _clean_schema_identifier(row.get("TABLE_NAME"))
                view_name = _clean_schema_identifier(row.get("VIEW_NAME"))
                column_name = _clean_schema_identifier(row.get("COLUMN_NAME"))
                src_table = _clean_schema_identifier(row.get("SRC_TABLE"))
                tgt_table = _clean_schema_identifier(row.get("TGT_TABLE"))
                src_column = _clean_schema_identifier(row.get("SRC_COLUMN"))
                tgt_column = _clean_schema_identifier(row.get("TGT_COLUMN"))
                fk_name = _clean_schema_identifier(row.get("FK_NAME") or row.get("CONSTRAINT_NAME"))

                if table_name:
                    tables.add(table_name)
                if view_name:
                    views.add(view_name)

                if table_name and column_name:
                    columns_by_object[table_name].add(column_name)
                elif view_name and column_name:
                    columns_by_object[view_name].add(column_name)

                if src_table:
                    tables.add(src_table)
                if tgt_table:
                    tables.add(tgt_table)
                if src_table and tgt_table and src_column and tgt_column:
                    fk_edges.add((src_table, src_column, tgt_table, tgt_column, fk_name or "FK"))

                text_sql = row.get("TEXT", "") or row.get("TEXT_VC", "")
                if view_name and text_sql:
                    for base_table in _extract_base_tables_from_view_sql(text_sql):
                        tables.add(base_table)
                        view_to_base_tables[view_name].add(base_table)

            source_stats[rel_path] = {"rows_seen": row_count, "status": "parsed"}
        except ValueError as exc:
            rejects.append(
                {
                    "source_path": rel_path,
                    "title": path.name,
                    "reason": str(exc) or "parse_error",
                    "asset_type": "sql_metadata",
                }
            )
            source_stats[rel_path] = {"rows_seen": row_count, "status": "parse_error", "error": str(exc)}
        except Exception as exc:
            rejects.append(
                {
                    "source_path": rel_path,
                    "title": path.name,
                    "reason": f"metadata_scan_exception:{exc.__class__.__name__}",
                    "asset_type": "sql_metadata",
                }
            )
            source_stats[rel_path] = {"rows_seen": row_count, "status": "parse_exception", "error": str(exc)}

    if not tables and not views and not fk_edges:
        return records, rejects, {
            "sources": [_relative_or_abs(path) for path in discovered_sources],
            "source_stats": source_stats,
            "note": "no_schema_signals_extracted",
        }

    table_modules: Dict[str, List[str]] = defaultdict(list)
    for table_name in sorted(tables):
        table_modules[_infer_module_from_table_name(table_name)].append(table_name)

    for module, module_tables in sorted(table_modules.items()):
        table_sample = module_tables[:180]
        rec = _manifest_record(
            source_path="backend/orawing_ai/data/crawler/metadata/metadata",
            source_uri=_oraclewings_uri(
                "backend/orawing_ai/data/crawler/metadata/metadata",
                anchor=f"tables_{_slug(module)}",
            ),
            title=f"Crawler Table Catalog {module}",
            module=module,
            task_type="sql_generation",
            doc_type=DocType.SCHEMA_METADATA,
            corpus=CorpusType.SCHEMA_METADATA,
            quality_score=0.88,
            content="\n".join(
                [
                    f"Module: {module}",
                    f"Detected table count: {len(module_tables)}",
                    "Sample tables:",
                    *[f"- {name}" for name in table_sample],
                ]
            ),
            trusted_schema_objects=module_tables[:1200],
            metadata={
                "metadata_source": "crawler_metadata_table_catalog",
                "module": module,
                "table_count": len(module_tables),
                "source_stats": source_stats,
            },
        )
        ok, reason = _validate_record(rec)
        if ok:
            records.append(rec)
        else:
            rejects.append(
                {
                    "source_path": rec["source_path"],
                    "title": rec["title"],
                    "reason": reason or "validation_failed",
                    "asset_type": "sql_metadata",
                }
            )

    if views:
        view_sample = sorted(views)[:220]
        rec = _manifest_record(
            source_path="backend/orawing_ai/data/crawler/metadata/metadata",
            source_uri=_oraclewings_uri(
                "backend/orawing_ai/data/crawler/metadata/metadata",
                anchor="view_catalog",
            ),
            title="Crawler View Catalog",
            module="Financials",
            task_type="sql_generation",
            doc_type=DocType.SCHEMA_METADATA,
            corpus=CorpusType.SCHEMA_METADATA,
            quality_score=0.86,
            content="\n".join(
                [
                    f"Detected view count: {len(views)}",
                    "Sample views:",
                    *[f"- {name}" for name in view_sample],
                ]
            ),
            trusted_schema_objects=sorted(views)[:1200],
            metadata={
                "metadata_source": "crawler_metadata_view_catalog",
                "view_count": len(views),
                "source_stats": source_stats,
            },
        )
        ok, reason = _validate_record(rec)
        if ok:
            records.append(rec)
        else:
            rejects.append(
                {
                    "source_path": rec["source_path"],
                    "title": rec["title"],
                    "reason": reason or "validation_failed",
                    "asset_type": "sql_metadata",
                }
            )

    # Keep record count controlled by selecting the highest-density objects.
    object_density = sorted(
        ((obj, len(cols)) for obj, cols in columns_by_object.items() if cols),
        key=lambda item: (-item[1], item[0]),
    )[:260]
    for object_name, col_count in object_density:
        columns = sorted(columns_by_object[object_name])[:220]
        module = _infer_module_from_table_name(object_name)
        rec = _manifest_record(
            source_path="backend/orawing_ai/data/crawler/metadata/metadata",
            source_uri=_oraclewings_uri(
                "backend/orawing_ai/data/crawler/metadata/metadata",
                anchor=f"columns_{_slug(object_name)}",
            ),
            title=f"Crawler Column Map {object_name}",
            module=module,
            task_type="sql_generation",
            doc_type=DocType.SCHEMA_METADATA,
            corpus=CorpusType.SCHEMA_METADATA,
            quality_score=0.87,
            content="\n".join(
                [
                    f"Object: {object_name}",
                    f"Detected columns: {col_count}",
                    "Columns:",
                    *[f"- {column}" for column in columns],
                ]
            ),
            trusted_schema_objects=[object_name] + columns[:280],
            metadata={
                "metadata_source": "crawler_metadata_column_map",
                "object_name": object_name,
                "column_count": col_count,
            },
        )
        ok, reason = _validate_record(rec)
        if ok:
            records.append(rec)
        else:
            rejects.append(
                {
                    "source_path": rec["source_path"],
                    "title": rec["title"],
                    "reason": reason or "validation_failed",
                    "asset_type": "sql_metadata",
                }
            )

    if fk_edges:
        edges_by_source: Dict[str, List[Tuple[str, str, str, str, str]]] = defaultdict(list)
        for edge in sorted(fk_edges):
            edges_by_source[edge[0]].append(edge)
        for source_table, source_edges in sorted(edges_by_source.items(), key=lambda item: (-len(item[1]), item[0]))[:220]:
            module = _infer_module_from_table_name(source_table)
            edge_lines = [
                f"- {src_table}.{src_col} -> {tgt_table}.{tgt_col} ({fk_name})"
                for src_table, src_col, tgt_table, tgt_col, fk_name in source_edges[:28]
            ]
            trusted_objects: List[str] = [source_table]
            for _, _, tgt_table, _, _ in source_edges[:28]:
                trusted_objects.append(tgt_table)
            rec = _manifest_record(
                source_path="backend/orawing_ai/data/crawler/metadata/metadata",
                source_uri=_oraclewings_uri(
                    "backend/orawing_ai/data/crawler/metadata/metadata",
                    anchor=f"fk_{_slug(source_table)}",
                ),
                title=f"Crawler FK Graph {source_table}",
                module=module,
                task_type="sql_generation",
                doc_type=DocType.SCHEMA_METADATA,
                corpus=CorpusType.SCHEMA_METADATA,
                quality_score=0.89,
                content="\n".join(
                    [
                        f"Source table: {source_table}",
                        f"Detected FK edges: {len(source_edges)}",
                        "Sample edges:",
                        *edge_lines,
                    ]
                ),
                trusted_schema_objects=sorted(set(trusted_objects)),
                metadata={
                    "metadata_source": "crawler_metadata_fk_graph",
                    "source_table": source_table,
                    "fk_edge_count": len(source_edges),
                },
            )
            ok, reason = _validate_record(rec)
            if ok:
                records.append(rec)
            else:
                rejects.append(
                    {
                        "source_path": rec["source_path"],
                        "title": rec["title"],
                        "reason": reason or "validation_failed",
                        "asset_type": "sql_metadata",
                    }
                )

    for view_name, base_tables in sorted(view_to_base_tables.items())[:220]:
        if not base_tables:
            continue
        module = _infer_module_from_table_name(view_name)
        base_sample = sorted(base_tables)[:32]
        rec = _manifest_record(
            source_path="backend/orawing_ai/data/crawler/metadata/metadata",
            source_uri=_oraclewings_uri(
                "backend/orawing_ai/data/crawler/metadata/metadata",
                anchor=f"viewexp_{_slug(view_name)}",
            ),
            title=f"Crawler View Expansion {view_name}",
            module=module,
            task_type="sql_generation",
            doc_type=DocType.SCHEMA_METADATA,
            corpus=CorpusType.SCHEMA_METADATA,
            quality_score=0.87,
            content="\n".join(
                [
                    f"View: {view_name}",
                    f"Detected base tables: {len(base_tables)}",
                    "Base tables:",
                    *[f"- {name}" for name in base_sample],
                ]
            ),
            trusted_schema_objects=[view_name] + base_sample,
            metadata={
                "metadata_source": "crawler_metadata_view_expansion",
                "view_name": view_name,
                "base_table_count": len(base_tables),
            },
        )
        ok, reason = _validate_record(rec)
        if ok:
            records.append(rec)
        else:
            rejects.append(
                {
                    "source_path": rec["source_path"],
                    "title": rec["title"],
                    "reason": reason or "validation_failed",
                    "asset_type": "sql_metadata",
                }
            )

    notes = {
        "sources": [_relative_or_abs(path) for path in discovered_sources],
        "source_stats": source_stats,
        "unique_tables_detected": len(tables),
        "unique_views_detected": len(views),
        "objects_with_column_maps": len(columns_by_object),
        "fk_edges_detected": len(fk_edges),
        "view_expansions_detected": len(view_to_base_tables),
    }
    return records, rejects, notes


def _sql_examples_full_scan_stage1() -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    rejects: List[Dict[str, Any]] = []
    seen_sql_hashes: set[str] = set()
    allowed_objects = _load_allowed_oracle_sql_objects()
    scanned_files = 0
    skipped_large_files = 0
    skipped_binary_or_unreadable = 0
    sql_candidates_seen = 0
    candidates_rejected_not_allowed = 0
    accepted_from_file: Counter[str] = Counter()

    for path in _iter_full_repo_scan_files(ORACLEWINGS_DIR):
        scanned_files += 1
        try:
            if path.stat().st_size > FULL_SQL_SCAN_MAX_FILE_BYTES:
                skipped_large_files += 1
                continue
        except OSError:
            skipped_binary_or_unreadable += 1
            continue

        try:
            text = path.read_text(encoding="utf-8", errors="ignore")
        except OSError:
            skipped_binary_or_unreadable += 1
            continue
        if not text or ("select" not in text.lower() and "from" not in text.lower()):
            continue

        rel_path = _relative_or_abs(path)
        sql_blocks = _split_candidate_sql_blocks(text)
        if not sql_blocks:
            continue

        for block_idx, sql_block in enumerate(sql_blocks, start=1):
            sql_candidates_seen += 1
            normalized_sql = sql_block.strip().rstrip(";")
            if not normalized_sql:
                continue
            reject_reason = _is_probably_unsupported_sql(normalized_sql)
            if reject_reason:
                rejects.append(
                    {
                        "source_path": rel_path,
                        "title": f"{path.name} SQL candidate #{block_idx}",
                        "reason": reject_reason,
                        "asset_type": "sql_examples",
                    }
                )
                continue

            table_names = _extract_tables_from_sql(normalized_sql)
            if not table_names:
                rejects.append(
                    {
                        "source_path": rel_path,
                        "title": f"{path.name} SQL candidate #{block_idx}",
                        "reason": "no_grounded_tables_detected",
                        "asset_type": "sql_examples",
                    }
                )
                continue
            trusted_tables = [name for name in table_names if name in allowed_objects]
            if not trusted_tables:
                candidates_rejected_not_allowed += 1
                rejects.append(
                    {
                        "source_path": rel_path,
                        "title": f"{path.name} SQL candidate #{block_idx}",
                        "reason": "no_allowed_oracle_objects_in_sql",
                        "asset_type": "sql_examples",
                    }
                )
                continue

            dedupe_key = stable_hash("sql_full_scan", _normalize(normalized_sql))
            if dedupe_key in seen_sql_hashes:
                continue
            seen_sql_hashes.add(dedupe_key)

            module = _infer_module_from_tables(trusted_tables)
            title = f"SQL Example {path.name} #{block_idx}"
            rec = _manifest_record(
                source_path=rel_path,
                source_uri=_oraclewings_uri(rel_path, anchor=f"sql_{block_idx}"),
                title=title,
                module=module,
                task_type="sql_generation",
                doc_type=DocType.SQL_EXAMPLE,
                corpus=CorpusType.SQL_EXAMPLES,
                quality_score=0.82,
                content=normalized_sql + ";",
                trusted_schema_objects=trusted_tables[:80],
                metadata={
                    "source_kind": "full_repo_sql_scan",
                    "source_extension": path.suffix.lower(),
                    "table_count": len(table_names),
                    "tables": trusted_tables[:80],
                    "raw_extracted_tables": table_names[:80],
                },
            )
            ok, reason = _validate_record(rec)
            if ok:
                records.append(rec)
                accepted_from_file[rel_path] += 1
            else:
                rejects.append(
                    {
                        "source_path": rel_path,
                        "title": title,
                        "reason": reason or "validation_failed",
                        "asset_type": "sql_examples",
                    }
                )
            if len(records) >= FULL_SQL_SCAN_MAX_ACCEPTED_RECORDS:
                break
        if len(records) >= FULL_SQL_SCAN_MAX_ACCEPTED_RECORDS:
            break

    notes = {
        "scan_root": str(ORACLEWINGS_DIR),
        "scanned_files": scanned_files,
        "skipped_large_files": skipped_large_files,
        "skipped_unreadable_files": skipped_binary_or_unreadable,
        "sql_candidates_seen": sql_candidates_seen,
        "candidates_rejected_not_allowed_oracle_objects": candidates_rejected_not_allowed,
        "allowed_oracle_object_count": len(allowed_objects),
        "accepted_records": len(records),
        "accepted_record_limit": FULL_SQL_SCAN_MAX_ACCEPTED_RECORDS,
        "top_sources": accepted_from_file.most_common(40),
    }
    return records, rejects, notes


def _sql_metadata_stage1() -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    rejects: List[Dict[str, Any]] = []

    modules_tables = _read_json(MODULES_TABLES_PATH)
    if isinstance(modules_tables, dict):
        for key, values in modules_tables.items():
            if not isinstance(values, list) or not values:
                continue
            module = _infer_module_from_modules_key(key)
            obj_kind = "views" if "views" in key.lower() else "tables"
            objects = [str(v).strip().replace('"', "") for v in values if str(v).strip()]
            trusted = objects[:800]
            content_lines = [
                f"Module key: {key}",
                f"Object kind: {obj_kind}",
                f"Object count: {len(objects)}",
                "Sample objects:",
                *[f"- {x}" for x in objects[:120]],
            ]
            rec = _manifest_record(
                source_path=str(MODULES_TABLES_PATH.relative_to(ORACLEWINGS_DIR)),
                source_uri=_oraclewings_uri(
                    str(MODULES_TABLES_PATH.relative_to(ORACLEWINGS_DIR)),
                    anchor=_slug(key),
                ),
                title=f"SQL Metadata {key}",
                module=module,
                task_type="sql_generation",
                doc_type=DocType.SCHEMA_METADATA,
                corpus=CorpusType.SCHEMA_METADATA,
                quality_score=0.9,
                content="\n".join(content_lines),
                trusted_schema_objects=trusted,
                metadata={
                    "metadata_source": "modules_tables",
                    "module_key": key,
                    "object_kind": obj_kind,
                    "object_count": len(objects),
                },
            )
            ok, reason = _validate_record(rec)
            if ok:
                records.append(rec)
            else:
                rejects.append(
                    {
                        "source_path": rec["source_path"],
                        "title": rec["title"],
                        "reason": reason or "validation_failed",
                        "asset_type": "sql_metadata",
                    }
                )

    join_hints = _read_json(JOIN_HINTS_PATH)
    hints = join_hints.get("hints", []) if isinstance(join_hints, dict) else []
    for idx, hint in enumerate(hints, start=1):
        source = str(hint.get("source", "")).strip()
        target = str(hint.get("target", "")).strip()
        condition = str(hint.get("condition", "")).strip()
        if not source or not target or not condition:
            rejects.append(
                {
                    "source_path": str(JOIN_HINTS_PATH.relative_to(ORACLEWINGS_DIR)),
                    "title": f"Join hint #{idx}",
                    "reason": "missing_source_target_or_condition",
                    "asset_type": "sql_metadata",
                }
            )
            continue
        module = _infer_module_from_table_name(source)
        rec = _manifest_record(
            source_path=str(JOIN_HINTS_PATH.relative_to(ORACLEWINGS_DIR)),
            source_uri=_oraclewings_uri(str(JOIN_HINTS_PATH.relative_to(ORACLEWINGS_DIR)), anchor=f"hint_{idx}"),
            title=f"Join Hint {source} -> {target}",
            module=module,
            task_type="sql_generation",
            doc_type=DocType.SCHEMA_METADATA,
            corpus=CorpusType.SCHEMA_METADATA,
            quality_score=0.88,
            content="\n".join(
                [
                    f"Join source: {source}",
                    f"Join target: {target}",
                    f"Join type: {hint.get('type', 'INNER')}",
                    f"Condition: {condition}",
                    f"Description: {_normalize(str(hint.get('description', '')))}",
                ]
            ),
            trusted_schema_objects=[source, target],
            metadata={
                "metadata_source": "join_hints",
                "join_type": hint.get("type", "INNER"),
                "condition": condition,
            },
        )
        ok, reason = _validate_record(rec)
        if ok:
            records.append(rec)
        else:
            rejects.append(
                {
                    "source_path": rec["source_path"],
                    "title": rec["title"],
                    "reason": reason or "validation_failed",
                    "asset_type": "sql_metadata",
                }
            )

    column_semantics = _read_json(COLUMN_SEMANTICS_PATH)
    if isinstance(column_semantics, dict):
        for semantic_group, cols in column_semantics.items():
            if not isinstance(cols, list) or not cols:
                continue
            objects = [str(c).strip().replace('"', "") for c in cols if str(c).strip()]
            rec = _manifest_record(
                source_path=str(COLUMN_SEMANTICS_PATH.relative_to(ORACLEWINGS_DIR)),
                source_uri=_oraclewings_uri(
                    str(COLUMN_SEMANTICS_PATH.relative_to(ORACLEWINGS_DIR)),
                    anchor=_slug(semantic_group),
                ),
                title=f"Column Semantics {semantic_group}",
                module="Financials",
                task_type="sql_generation",
                doc_type=DocType.SCHEMA_METADATA,
                corpus=CorpusType.SCHEMA_METADATA,
                quality_score=0.86,
                content="\n".join([f"Semantic group: {semantic_group}", *[f"- {c}" for c in objects[:160]]]),
                trusted_schema_objects=objects[:800],
                metadata={
                    "metadata_source": "column_semantics",
                    "semantic_group": semantic_group,
                    "column_count": len(objects),
                },
            )
            ok, reason = _validate_record(rec)
            if ok:
                records.append(rec)
            else:
                rejects.append(
                    {
                        "source_path": rec["source_path"],
                        "title": rec["title"],
                        "reason": reason or "validation_failed",
                        "asset_type": "sql_metadata",
                    }
                )

    view_expansions = _read_json(VIEW_EXPANSIONS_PATH)
    if isinstance(view_expansions, dict):
        for view_name, tables in view_expansions.items():
            if isinstance(tables, list):
                base_tables = [str(t).strip().replace('"', "") for t in tables if str(t).strip()]
            elif isinstance(tables, dict):
                base_tables = [str(t).strip().replace('"', "") for t in tables.get("base_tables", []) if str(t).strip()]
            else:
                base_tables = []
            if not base_tables:
                rejects.append(
                    {
                        "source_path": str(VIEW_EXPANSIONS_PATH.relative_to(ORACLEWINGS_DIR)),
                        "title": f"View expansion {view_name}",
                        "reason": "missing_base_tables",
                        "asset_type": "sql_metadata",
                    }
                )
                continue
            module = _infer_module_from_table_name(str(view_name))
            rec = _manifest_record(
                source_path=str(VIEW_EXPANSIONS_PATH.relative_to(ORACLEWINGS_DIR)),
                source_uri=_oraclewings_uri(
                    str(VIEW_EXPANSIONS_PATH.relative_to(ORACLEWINGS_DIR)),
                    anchor=_slug(str(view_name)),
                ),
                title=f"View Expansion {view_name}",
                module=module,
                task_type="sql_generation",
                doc_type=DocType.SCHEMA_METADATA,
                corpus=CorpusType.SCHEMA_METADATA,
                quality_score=0.89,
                content="\n".join(
                    [
                        f"View: {view_name}",
                        "Base tables:",
                        *[f"- {t}" for t in base_tables],
                    ]
                ),
                trusted_schema_objects=[str(view_name)] + base_tables,
                metadata={
                    "metadata_source": "view_expansions",
                    "view_name": view_name,
                    "base_table_count": len(base_tables),
                },
            )
            ok, reason = _validate_record(rec)
            if ok:
                records.append(rec)
            else:
                rejects.append(
                    {
                        "source_path": rec["source_path"],
                        "title": rec["title"],
                        "reason": reason or "validation_failed",
                        "asset_type": "sql_metadata",
                    }
                )

    # DDL summaries
    for ddl_path in DDL_FILES:
        if not ddl_path.exists():
            rejects.append(
                {
                    "source_path": str(ddl_path),
                    "title": ddl_path.name,
                    "reason": "ddl_file_missing",
                    "asset_type": "sql_metadata",
                }
            )
            continue
        tables, views = _parse_ddl_objects(ddl_path)
        trusted_objects = (tables + views)[:1200]
        if not trusted_objects:
            rejects.append(
                {
                    "source_path": str(ddl_path.relative_to(ORACLEWINGS_DIR)),
                    "title": ddl_path.name,
                    "reason": "no_tables_or_views_detected",
                    "asset_type": "sql_metadata",
                }
            )
            continue
        name_lower = ddl_path.name.lower()
        if "hcm" in name_lower:
            module = "HCM"
        elif "procurement" in name_lower:
            module = "Procurement"
        elif "scm" in name_lower:
            module = "SCM"
        elif "project" in name_lower:
            module = "Projects"
        else:
            module = "Financials"

        content_lines = [
            f"DDL file: {ddl_path.name}",
            f"Detected tables: {len(tables)}",
            f"Detected views: {len(views)}",
            "Sample tables:",
            *[f"- {t}" for t in tables[:140]],
            "Sample views:",
            *[f"- {v}" for v in views[:120]],
        ]
        rec = _manifest_record(
            source_path=str(ddl_path.relative_to(ORACLEWINGS_DIR)),
            source_uri=_oraclewings_uri(str(ddl_path.relative_to(ORACLEWINGS_DIR))),
            title=f"DDL Metadata {ddl_path.name}",
            module=module,
            task_type="sql_generation",
            doc_type=DocType.SCHEMA_METADATA,
            corpus=CorpusType.SCHEMA_METADATA,
            quality_score=0.84,
            content="\n".join(content_lines),
            trusted_schema_objects=trusted_objects,
            metadata={
                "metadata_source": "ddl_summary",
                "table_count": len(tables),
                "view_count": len(views),
                "excluded_for_now": False,
            },
        )
        ok, reason = _validate_record(rec)
        if ok:
            records.append(rec)
        else:
            rejects.append(
                {
                    "source_path": rec["source_path"],
                    "title": rec["title"],
                    "reason": reason or "validation_failed",
                    "asset_type": "sql_metadata",
                }
            )

    crawler_records, crawler_rejects, crawler_notes = _sql_metadata_crawler_stage1()
    records.extend(crawler_records)
    rejects.extend(crawler_rejects)

    notes = {
        "sources": [
            str(MODULES_TABLES_PATH.relative_to(ORACLEWINGS_DIR)),
            str(JOIN_HINTS_PATH.relative_to(ORACLEWINGS_DIR)),
            str(COLUMN_SEMANTICS_PATH.relative_to(ORACLEWINGS_DIR)),
            str(VIEW_EXPANSIONS_PATH.relative_to(ORACLEWINGS_DIR)),
            *[str(p.relative_to(ORACLEWINGS_DIR)) for p in DDL_FILES],
        ],
        "crawler_metadata": crawler_notes,
    }
    return records, rejects, notes


def _troubleshooting_stage1() -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    rejects: List[Dict[str, Any]] = []

    data = _read_json(TROUBLESHOOTING_KB_PATH)
    if not isinstance(data, dict):
        return records, [
            {
                "source_path": str(TROUBLESHOOTING_KB_PATH.relative_to(ORACLEWINGS_DIR)),
                "title": "troubleshooting_kb",
                "reason": "invalid_json_structure",
                "asset_type": "troubleshooting",
            }
        ], {"sources": [str(TROUBLESHOOTING_KB_PATH.relative_to(ORACLEWINGS_DIR))]}

    def add_record(
        *,
        title: str,
        symptom: str,
        causes: List[str],
        fixes: List[str],
        anchor: str,
        quality: float = 0.9,
    ) -> None:
        if not symptom or len(_normalize(symptom)) < 8:
            rejects.append(
                {
                    "source_path": str(TROUBLESHOOTING_KB_PATH.relative_to(ORACLEWINGS_DIR)),
                    "title": title,
                    "reason": "symptom_missing_or_too_short",
                    "asset_type": "troubleshooting",
                }
            )
            return
        if not causes:
            rejects.append(
                {
                    "source_path": str(TROUBLESHOOTING_KB_PATH.relative_to(ORACLEWINGS_DIR)),
                    "title": title,
                    "reason": "cause_missing_strict_filter",
                    "asset_type": "troubleshooting",
                }
            )
            return
        if not fixes:
            rejects.append(
                {
                    "source_path": str(TROUBLESHOOTING_KB_PATH.relative_to(ORACLEWINGS_DIR)),
                    "title": title,
                    "reason": "fix_steps_missing_strict_filter",
                    "asset_type": "troubleshooting",
                }
            )
            return

        content = "\n".join(
            [
                f"Symptom: {symptom}",
                "Likely Causes:",
                *[f"- {c}" for c in causes],
                "Resolution Steps:",
                *[f"{idx}. {step}" for idx, step in enumerate(fixes, start=1)],
            ]
        )
        rec = _manifest_record(
            source_path=str(TROUBLESHOOTING_KB_PATH.relative_to(ORACLEWINGS_DIR)),
            source_uri=_oraclewings_uri(str(TROUBLESHOOTING_KB_PATH.relative_to(ORACLEWINGS_DIR)), anchor=anchor),
            title=title,
            module="Procurement",
            task_type="troubleshooting",
            doc_type=DocType.TROUBLESHOOTING_DOC,
            corpus=CorpusType.TROUBLESHOOTING,
            quality_score=quality,
            content=content,
            trusted_schema_objects=[],
            metadata={
                "symptom": symptom,
                "probable_causes": causes,
                "resolution_steps": fixes,
                "strict_filter_passed": True,
                "source_section": anchor,
            },
        )
        ok, reason = _validate_record(rec)
        if ok:
            records.append(rec)
        else:
            rejects.append(
                {
                    "source_path": rec["source_path"],
                    "title": rec["title"],
                    "reason": reason or "validation_failed",
                    "asset_type": "troubleshooting",
                }
            )

    for idx, row in enumerate(data.get("fusion_errors", []), start=1):
        add_record(
            title=f"SCM Troubleshooting Fusion Error {idx}",
            symptom=str(row.get("error", "")).strip(),
            causes=[str(x).strip() for x in row.get("causes", []) if str(x).strip()],
            fixes=[str(x).strip() for x in row.get("solutions", []) if str(x).strip()],
            anchor=f"fusion_errors[{idx}]",
            quality=0.92,
        )

    for idx, row in enumerate(data.get("ora_errors", []), start=1):
        meaning = str(row.get("meaning", "")).strip()
        causes = [meaning] if meaning else []
        add_record(
            title=f"SCM Troubleshooting ORA Error {idx}",
            symptom=str(row.get("error", "")).strip(),
            causes=causes,
            fixes=[str(x).strip() for x in row.get("solutions", []) if str(x).strip()],
            anchor=f"ora_errors[{idx}]",
            quality=0.9,
        )

    for idx, row in enumerate(data.get("workflow_errors", []), start=1):
        add_record(
            title=f"SCM Troubleshooting Workflow Error {idx}",
            symptom=str(row.get("error", "")).strip(),
            causes=[str(x).strip() for x in row.get("causes", []) if str(x).strip()],
            fixes=[str(x).strip() for x in row.get("solutions", []) if str(x).strip()],
            anchor=f"workflow_errors[{idx}]",
            quality=0.9,
        )

    # common_issues are rejected unless cause exists (strict symptom->cause->fix policy)
    for idx, row in enumerate(data.get("common_issues", []), start=1):
        title = f"SCM Troubleshooting Common Issue {idx}"
        symptom = str(row.get("issue", "")).strip()
        fixes = [str(x).strip() for x in row.get("solutions", []) if str(x).strip()]
        add_record(
            title=title,
            symptom=symptom,
            causes=[],  # strict filter
            fixes=fixes,
            anchor=f"common_issues[{idx}]",
            quality=0.75,
        )

    notes = {
        "sources": [str(TROUBLESHOOTING_KB_PATH.relative_to(ORACLEWINGS_DIR))],
        "strict_filter": "symptom->cause->fix completeness required",
    }
    return records, rejects, notes


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=True), encoding="utf-8")


def _write_jsonl(path: Path, rows: List[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=True) + "\n")


def _index_manifest_rows(manifest_rows: List[Dict[str, Any]], tenant: str) -> Dict[str, Any]:
    indexed = 0
    by_corpus = Counter()
    faiss_cache: Dict[str, FaissIndex] = {}
    for payload in manifest_rows:
        record = IngestionManifestRecord(**payload)
        ok, reason = CuratedIngestionValidator.validate_manifest_record(record)
        if not ok:
            continue
        if not record.content:
            continue

        document = CuratedIngestionValidator.build_document(
            source_path=record.source_path,
            source_uri=record.metadata.get("source_uri", record.source_path),
            title=record.title,
            module=record.module,
            task_type=record.task_type,
            doc_type=record.doc_type,
            trusted_schema_objects=record.trusted_schema_objects,
            quality_score=record.quality_score,
            source_system=record.source_system,
            content=record.content,
            metadata=record.metadata,
        )
        chunk = CuratedIngestionValidator.build_chunk(document, document.content, 0)
        chunk_payload = CuratedIngestionValidator.chunk_payload(chunk)
        corpus = chunk_payload["metadata"]["corpus"]
        if corpus not in faiss_cache:
            faiss_cache[corpus] = FaissIndex(
                tenant_id=tenant,
                indexes_dir=str(INDEXES_DIR),
                corpus=corpus,
            )
        faiss_cache[corpus].add_chunks_list([chunk_payload], batch_size=1)
        indexed += 1
        by_corpus[corpus] += 1
    return {"indexed_rows": indexed, "indexed_by_corpus": dict(by_corpus)}


def _filter_excluded(records: List[Dict[str, Any]], rejects: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    kept: List[Dict[str, Any]] = []
    for rec in records:
        source_path = rec.get("source_path", "").lower()
        if any(hint in source_path for hint in EXCLUDED_SOURCE_HINTS):
            rejects.append(
                {
                    "source_path": rec.get("source_path", ""),
                    "title": rec.get("title", ""),
                    "reason": "excluded_by_plan_policy",
                    "asset_type": "policy_exclusion",
                }
            )
        else:
            kept.append(rec)
    return kept


def run_stage1(tenant: str, index: bool, full_sql_scan: bool = False) -> Dict[str, Any]:
    generated_at = datetime.now(timezone.utc).isoformat()
    context = BuildContext(generated_at_utc=generated_at, tenant=tenant, index_enabled=index)

    ff_records, ff_rejects, ff_notes = _fast_formula_stage1()
    sql_records, sql_rejects, sql_notes = _sql_metadata_stage1()
    sql_example_records: List[Dict[str, Any]] = []
    sql_example_rejects: List[Dict[str, Any]] = []
    sql_example_notes: Dict[str, Any] = {}
    if full_sql_scan:
        sql_example_records, sql_example_rejects, sql_example_notes = _sql_examples_full_scan_stage1()
    trb_records, trb_rejects, trb_notes = _troubleshooting_stage1()

    all_rejects = ff_rejects + sql_rejects + sql_example_rejects + trb_rejects
    ff_records = _filter_excluded(ff_records, all_rejects)
    sql_records = _filter_excluded(sql_records, all_rejects)
    sql_example_records = _filter_excluded(sql_example_records, all_rejects)
    trb_records = _filter_excluded(trb_records, all_rejects)

    ff_manifest = MANIFEST_DIR / "stage1_fast_formula_manifest.jsonl"
    sql_manifest = MANIFEST_DIR / "stage1_schema_metadata_manifest.jsonl"
    sql_examples_manifest = MANIFEST_DIR / "stage1_sql_examples_manifest.jsonl"
    trb_manifest = MANIFEST_DIR / "stage1_troubleshooting_manifest.jsonl"
    reject_manifest = MANIFEST_DIR / "stage1_reject_manifest.jsonl"

    _write_jsonl(ff_manifest, ff_records)
    _write_jsonl(sql_manifest, sql_records)
    _write_jsonl(sql_examples_manifest, sql_example_records)
    _write_jsonl(trb_manifest, trb_records)
    _write_jsonl(reject_manifest, all_rejects)

    indexed_stats = {"indexed_rows": 0, "indexed_by_corpus": {}}
    if index:
        indexed_stats = _index_manifest_rows(
            ff_records + sql_records + sql_example_records + trb_records,
            tenant=tenant,
        )

    summary = {
        "generated_at_utc": context.generated_at_utc,
        "tenant": context.tenant,
        "index_enabled": context.index_enabled,
        "source_root": str(ORACLEWINGS_DIR),
        "plan_alignment": {
            "stage_1_fast_formula_structured_assets": True,
            "stage_1_sql_metadata_assets": True,
            "stage_1_troubleshooting_kb_strict_filter": True,
            "stage_1_full_repo_sql_scan_enabled": full_sql_scan,
            "stage_2_deferred": True,
            "stage_3_deferred": True,
        },
        "excluded_policy": [
            ".db/.sqlite/.pyc/.pack/.idx excluded",
            "generated temp snippets excluded",
            "direct BI Publisher parameterized SQL excluded",
            "frontend/test artifacts deferred",
        ],
        "manifest_paths": {
            "fast_formula": str(ff_manifest),
            "sql_metadata": str(sql_manifest),
            "sql_examples": str(sql_examples_manifest),
            "troubleshooting": str(trb_manifest),
            "rejects": str(reject_manifest),
        },
        "counts": {
            "fast_formula_records": len(ff_records),
            "sql_metadata_records": len(sql_records),
            "sql_example_records": len(sql_example_records),
            "troubleshooting_records": len(trb_records),
            "reject_records": len(all_rejects),
            "total_stage1_records": len(ff_records) + len(sql_records) + len(sql_example_records) + len(trb_records),
        },
        "notes": {
            "fast_formula": ff_notes,
            "sql_metadata": sql_notes,
            "sql_examples_full_scan": sql_example_notes,
            "troubleshooting": trb_notes,
        },
        "index_stats": indexed_stats,
        "next_benchmark_sequence": [
            "Fast Formula targeted benchmark (generation + troubleshooting + refusal cases)",
            "SQL targeted benchmark (field/join/filter/style/verifier gates)",
            "Troubleshooting benchmark for modules with new KB coverage",
            "200-case mixed regression slice",
            "1000/5000 broad refresh only after targeted thresholds pass",
        ],
    }

    _write_json(OUTPUT_DIR / "stage1_execution_summary.json", summary)
    (OUTPUT_DIR / "stage1_execution_report.md").write_text(
        "\n".join(
            [
                "# Oraclewings Post-Freeze Stage 1 Execution Report",
                "",
                f"- Generated at (UTC): `{summary['generated_at_utc']}`",
                f"- Tenant: `{tenant}`",
                f"- Indexing enabled: `{index}`",
                "",
                "## Record Counts",
                f"- Fast Formula records: `{len(ff_records)}`",
                f"- SQL metadata records: `{len(sql_records)}`",
                f"- SQL example records: `{len(sql_example_records)}`",
                f"- Troubleshooting records: `{len(trb_records)}`",
                f"- Reject records: `{len(all_rejects)}`",
                "",
                "## Manifest Paths",
                f"- `{ff_manifest}`",
                f"- `{sql_manifest}`",
                f"- `{sql_examples_manifest}`",
                f"- `{trb_manifest}`",
                f"- `{reject_manifest}`",
                "",
                "## Next Step",
                "- Run targeted benchmarks in the sequence listed in `stage1_execution_summary.json`.",
            ]
        ),
        encoding="utf-8",
    )
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description="Execute Oraclewings post-freeze Stage-1 ingestion plan.")
    parser.add_argument("--tenant", default="oraclewings_stage1", help="Tenant ID for optional indexing.")
    parser.add_argument("--index", action="store_true", help="Index stage-1 manifest rows into Faiss corpora.")
    parser.add_argument(
        "--full-sql-scan",
        action="store_true",
        help="Recursively scan full Oraclewings repo for SQL example candidates.",
    )
    args = parser.parse_args()

    summary = run_stage1(tenant=args.tenant, index=args.index, full_sql_scan=args.full_sql_scan)
    print(json.dumps(summary["counts"], indent=2))
    if args.index:
        print(json.dumps(summary["index_stats"], indent=2))


if __name__ == "__main__":
    main()
