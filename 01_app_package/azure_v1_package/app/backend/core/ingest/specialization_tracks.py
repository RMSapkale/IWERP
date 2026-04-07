import ast
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import sqlglot
from sqlglot import exp

from core.grounding.trusted_registry import get_default_registry
from core.grounding.verifier import Verifier
from core.ingest.curation import CuratedIngestionValidator, stable_hash
from core.schemas.curation import DocType, SourceSystem
from core.schemas.router import FusionModule, TaskType


ROOT_DIR = Path("/Users/integrationwings/Desktop/LLM_Wrap")
BASE_DIR = ROOT_DIR / "iwerp-prod"
INDEXES_DIR = BASE_DIR / "backend" / "core" / "retrieval" / "vectors"
SPECIALIZATION_DIR = BASE_DIR / "specialization_tracks"
MANIFEST_DIR = SPECIALIZATION_DIR / "manifests"

ORG_SQL_DIR = ROOT_DIR / "oracle-fusion-slm" / "oracle-fusion-slm" / "data" / "source_raw" / "organized_sql"
ORACLEWINGS_FUSION_QUERIES_PATH = (
    ROOT_DIR
    / "Oraclewings_ai"
    / "backend"
    / "orawing_ai"
    / "examples"
    / "fusion_queries.json"
)
FORMULA_KNOWLEDGE_PATH = ROOT_DIR / "Data_Complete" / "HCM" / "jsons" / "FAST_FORMULA_KNOWLEDGE.json"
FORMULA_CSV_PATH = ROOT_DIR / "Data_Complete" / "HCM" / "csvs" / "Fastformula_KT.csv"
FORMULA_TYPES_PATH = ROOT_DIR / "Data_Complete" / "Finance" / "jsons" / "formula_types.json"
FORMULA_SUPPORTING_JSON_PATHS = [
    ROOT_DIR / "Data_Complete" / "HCM" / "jsons" / "HCM_EXTRACT_KNOWLEDGE.json",
    ROOT_DIR / "Data_Complete" / "HCM" / "jsons" / "HCM_ORACLE_WEB_DATA.json",
]
FORMULA_SUPPORTING_TEXT_PATHS = [
    ROOT_DIR / "Data_Complete" / "HCM" / "csvs" / "HCM_Extract_Oracle_info.csv",
    ROOT_DIR / "oracle-fusion-slm" / "oracle-fusion-slm" / "data" / "source_raw" / "organized_sql"
    / "Human Capital Management"
    / "Workforce Management"
    / "Person Management"
    / "Data Models"
    / "Formula Data.xdm.txt",
]
HARVEST_MANIFEST_DIR = SPECIALIZATION_DIR / "oraclewings_ai_main_harvest" / "manifests"
HARVEST_FORMULA_EXAMPLES_PATH = HARVEST_MANIFEST_DIR / "fast_formula_examples.jsonl"
HARVEST_FORMULA_SUPPORT_PATH = HARVEST_MANIFEST_DIR / "fast_formula_supporting_docs.jsonl"
HARVEST_HDL_FBDI_PATH = HARVEST_MANIFEST_DIR / "hdl_fbdi_knowledge_corpus.jsonl"
HARVEST_SQL_RUNTIME_PATH = HARVEST_MANIFEST_DIR / "sql_runtime_assets.jsonl"

PATH_MODULE_HINTS: List[Tuple[re.Pattern[str], str]] = [
    (re.compile(r"(?i)(accounts payable|\bAP\b|payables|supplier)"), FusionModule.PAYABLES.value),
    (re.compile(r"(?i)(accounts receivable|\bAR\b|receivables|receipt|customer)"), FusionModule.RECEIVABLES.value),
    (re.compile(r"(?i)(general ledger|\bGL\b|journal|ledger)"), FusionModule.GENERAL_LEDGER.value),
    (re.compile(r"(?i)(cash management|bank|reconciliation|\bCE\b)"), FusionModule.CASH_MANAGEMENT.value),
    (re.compile(r"(?i)(asset|depreciation|\bFA\b)"), FusionModule.ASSETS.value),
    (re.compile(r"(?i)(procurement|purchasing|supplier portal|requisition|\bPO\b|\bPOZ\b|\bPOR\b)"), FusionModule.PROCUREMENT.value),
    (re.compile(r"(?i)(inventory|shipping|order management|manufacturing|planning|supply chain|\bINV\b|\bDOO\b|\bMSC\b|\bWSH\b)"), FusionModule.SCM.value),
    (re.compile(r"(?i)(hcm|payroll|benefit|absence|employee|fast formula|\bPER\b|\bPAY\b)"), FusionModule.HCM.value),
    (re.compile(r"(?i)(projects|ppm|grant|\bPJT\b|\bPJC\b|\bPJF\b)"), FusionModule.PROJECTS.value),
    (re.compile(r"(?i)(tax|\bZX\b)"), FusionModule.TAX.value),
]
ORACLEWINGS_SQL_TABLE_REMAP: Dict[str, str] = {
    "DOO_HEADER_ALL": "DOO_HEADERS_ALL",
    "FA_ADDITIONS": "FA_ADDITIONS_B",
    "FA_CATEGORIES": "FA_CATEGORIES_B",
    "HZ_CUST_ACCOUNTS": "HZ_CUST_ACCOUNTS_",
    "PJF_TASKS_VL": "PJF_TASKS_V",
    "PER_ABSENCE_TYPES_VL": "ANC_ABSENCE_TYPES_VL",
    "PO_REQUISITION_HEADERS_ALL": "POR_REQUISITION_HEADERS_ALL",
    "PO_REQUISITION_LINES_ALL": "POR_REQUISITION_LINES_ALL",
}

FORMULA_FUNCTIONS = {
    "ABS",
    "ADD_DAYS",
    "ADD_MONTHS",
    "CEIL",
    "DAYS_BETWEEN",
    "FLOOR",
    "GREATEST",
    "INSTR",
    "LEAST",
    "LENGTH",
    "MOD",
    "MONTHS_BETWEEN",
    "POWER",
    "ROUND",
    "SQRT",
    "SUBSTR",
    "TO_CHAR",
    "TO_DATE",
    "TRUNC",
}
FORMULA_RESERVED = {
    "DEFAULT",
    "FOR",
    "IS",
    "INPUTS",
    "ARE",
    "IF",
    "THEN",
    "ELSE",
    "ENDIF",
    "END",
    "WHILE",
    "LOOP",
    "ENDLOOP",
    "RETURN",
    "AND",
    "OR",
    "NOT",
    "NULL",
}
AUDIT_COLUMN_PREFIXES = (
    "CREATED_",
    "CREATION_",
    "LAST_UPDATE_",
    "LAST_UPDATED_",
    "OBJECT_VERSION_",
    "REQUEST_ID",
    "PROGRAM_",
    "ATTRIBUTE",
    "AUDIT_",
    "ORA_SEED_",
)


def _read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=True), encoding="utf-8")


def _write_jsonl(path: Path, rows: Iterable[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=True) + "\n")


def _clean_title(value: str) -> str:
    text = re.sub(r"\s+", " ", value.replace("_", " ").replace("-", " ")).strip()
    return text or "Untitled"


def _formula_type_from_text(text: str, known_types: List[str]) -> str:
    lowered = text.lower()
    best = "UNKNOWN"
    for formula_type in known_types:
        if formula_type.lower() in lowered:
            return formula_type
    heuristics = [
        ("proration", "Proration"),
        ("accrual", "Accrual"),
        ("absence", "Absence"),
        ("eligibility", "Participation and Rate Eligibility"),
        ("rate", "Rate Value Calculation"),
        ("payroll", "Oracle Payroll"),
        ("compensation", "Compensation Calculation"),
        ("balance", "Balance Adjustment"),
        ("validation", "Global Absence Entry Validation"),
        ("time", "Time Calculation Rules"),
    ]
    for token, label in heuristics:
        if token in lowered:
            best = label
            break
    return best


def _extract_formula_name(title: str, content: str) -> str:
    match = re.search(r"(?im)^FORMULA NAME:\s*(.+)$", content)
    if match:
        return match.group(1).strip()
    return _clean_title(title)


def _extract_formula_inputs(content: str) -> List[str]:
    match = re.search(r"(?im)INPUTS\s+ARE\s+(.+)$", content)
    if not match:
        return []
    inputs = match.group(1).strip().strip('"')
    return [part.strip() for part in re.split(r"\s*,\s*", inputs) if part.strip()]


def _extract_formula_functions(content: str) -> List[str]:
    functions = {
        token.upper()
        for token in re.findall(r"\b([A-Z][A-Z0-9_]+)\s*\(", content)
        if token.upper() in FORMULA_FUNCTIONS
    }
    return sorted(functions)


def _extract_formula_database_items(content: str) -> List[str]:
    candidates = {
        token.upper()
        for token in re.findall(r"\b([A-Z][A-Z0-9_]{3,})\b", content)
        if token.upper() not in FORMULA_FUNCTIONS and token.upper() not in FORMULA_RESERVED
    }
    return sorted(candidates)


def _extract_formula_contexts(content: str) -> List[str]:
    contexts = set()
    for token in ("CALC_START_DATE", "CALC_END_DATE", "DATE_EARNED", "PAYROLL_REL_ACTION_ID"):
        if token in content.upper():
            contexts.add(token)
    for match in re.findall(r"(?im)^CONTEXTS?\s+ARE\s+(.+)$", content):
        contexts.update(part.strip() for part in match.split(",") if part.strip())
    return sorted(contexts)


def _extract_formula_output_pattern(content: str) -> str:
    match = re.search(r"(?im)^RETURN\s+(.+)$", content)
    return match.group(1).strip() if match else ""


def _extract_formula_defaults(content: str) -> List[str]:
    defaults = []
    for line in content.splitlines():
        stripped = line.strip().strip('"')
        if stripped.upper().startswith("DEFAULT FOR "):
            defaults.append(stripped)
    return defaults


def _extract_formula_use_case(title: str, content: str) -> str:
    match = re.search(r"(?im)^FORMULA NAME:\s*(.+)$", content)
    if match:
        return _clean_title(match.group(1))
    return _clean_title(title)


def _normalize_formula_line(line: str) -> str:
    stripped = line.strip()
    if stripped.startswith('"') and stripped.endswith('"') and len(stripped) >= 2:
        stripped = stripped[1:-1]
    return stripped


def _normalize_formula_record(
    record: Dict[str, Any],
    *,
    source_file: str,
    source_sheet: str = "",
    confidence: float,
) -> Dict[str, Any]:
    content = "\n".join(
        line for line in (_normalize_formula_line(line) for line in str(record["content"]).splitlines()) if line
    ).strip()
    formula_name = _extract_formula_name(str(record.get("title") or ""), content)
    formula_type = str(record.get("formula_type") or "UNKNOWN")
    use_case = str(record.get("use_case") or _extract_formula_use_case(str(record.get("title") or ""), content))
    input_values = _extract_formula_inputs(content)
    database_items = _extract_formula_database_items(content)
    contexts = _extract_formula_contexts(content)
    functions = _extract_formula_functions(content)
    output_pattern = _extract_formula_output_pattern(content)
    defaults = _extract_formula_defaults(content)
    normalized = dict(record)
    normalized.update(
        {
            "content": content,
            "formula_name": formula_name,
            "formula_type": formula_type,
            "use_case": use_case,
            "input_values": input_values,
            "contexts": contexts,
            "database_items": database_items,
            "functions": functions,
            "return_behavior": output_pattern,
            "default_handling": defaults,
            "source_file": source_file,
            "source_sheet": source_sheet,
            "confidence": round(float(confidence), 2),
        }
    )
    return normalized


def _read_jsonl(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        return []
    rows: List[Dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            rows.append(json.loads(line))
    return rows


def _candidate_formula_workbooks() -> List[Path]:
    candidates: List[Path] = []
    extensions = ("*.xlsx", "*.xlsm", "*.xls")
    name_tokens = ("formula", "fastformula", "fast_formula", "ff_", "ffsample", "sample_report")
    for pattern in extensions:
        for path in ROOT_DIR.rglob(pattern):
            if "node_modules" in path.parts:
                continue
            lowered = path.name.lower()
            if any(token in lowered for token in name_tokens):
                candidates.append(path)
    return sorted(set(candidates))


def _parse_loose_literal(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text or text[0] not in "[{(":
        return value
    try:
        return ast.literal_eval(text)
    except Exception:
        return value


def _harvest_preview_text(row: Dict[str, Any]) -> str:
    payload = row.get("normalized_payload") or {}
    parts = [
        str(row.get("title") or ""),
        str(payload.get("content_preview") or ""),
        str(payload.get("description") or ""),
        str(payload.get("formula_type") or ""),
        str(payload.get("section_title") or ""),
        str(payload.get("oracle_job") or ""),
        str(payload.get("table_name") or ""),
    ]
    return "\n".join(part for part in parts if part).strip()


def _normalize_harvest_formula_type(raw_value: Any) -> Tuple[str, str]:
    parsed = _parse_loose_literal(raw_value)
    if isinstance(parsed, dict):
        name = str(parsed.get("name") or parsed.get("formula_type") or "UNKNOWN").strip()
        description = str(parsed.get("description") or "").strip()
        return name or "UNKNOWN", description
    text = str(raw_value or "").strip()
    return text or "UNKNOWN", ""


def _sanitize_harvest_table_name(raw_value: Any) -> str:
    text = str(raw_value or "").strip().upper()
    if not text:
        return ""
    matches = re.findall(r"TABLE\s+NAME\s*:\s*([A-Z][A-Z0-9_$]{2,})", text)
    if matches:
        return matches[-1]
    tokens = [
        token
        for token in re.findall(r"\b[A-Z][A-Z0-9_$]{2,}\b", text)
        if token not in {"FILE", "LINK", "ASSET", "TABLE", "NAME", "NA"}
    ]
    if len(tokens) == 1:
        return tokens[0]
    underscore_tokens = [token for token in tokens if "_" in token]
    if underscore_tokens:
        return underscore_tokens[-1]
    return tokens[-1] if tokens else ""


def _formula_template_kind(formula_type: str, use_case: str) -> str:
    lowered = f"{formula_type} {use_case}".lower()
    if any(token in lowered for token in ("proration", "pro-ration")):
        return "proration"
    if "accrual" in lowered:
        return "accrual"
    if any(token in lowered for token in ("validation", "eligibility", "ineligibility")):
        return "validation"
    if any(token in lowered for token in ("extract", "record", "rule", "loader")):
        return "extract"
    if any(token in lowered for token in ("rate", "conversion")):
        return "rate"
    if any(token in lowered for token in ("balance", "compensation", "calculation", "payroll", "indirect")):
        return "payroll"
    if any(token in lowered for token in ("time", "workforce_management")):
        return "time"
    return "generic"


def _default_formula_traits(kind: str) -> Dict[str, Any]:
    defaults = {
        "proration": {
            "inputs": ["start_date (DATE)", "end_date (DATE)"],
            "db_items": ["CALC_START_DATE", "CALC_END_DATE"],
            "contexts": ["CALC_START_DATE", "CALC_END_DATE"],
            "functions": ["DAYS_BETWEEN"],
            "return_var": "proration_factor",
        },
        "accrual": {
            "inputs": ["hire_date (DATE)", "term_date (DATE)"],
            "db_items": ["PER_ASG_EFFECTIVE_START_DATE", "PAYROLL_REL_ACTION_ID"],
            "contexts": ["PAYROLL_REL_ACTION_ID", "DATE_EARNED"],
            "functions": ["MONTHS_BETWEEN", "ROUND"],
            "return_var": "accrual_value",
        },
        "validation": {
            "inputs": ["entry_value (TEXT)"],
            "db_items": ["PER_ASG_PERSON_ID", "PAYROLL_REL_ACTION_ID"],
            "contexts": ["PAYROLL_REL_ACTION_ID"],
            "functions": [],
            "return_var": "validation_result",
        },
        "extract": {
            "inputs": ["effective_date (DATE)"],
            "db_items": ["PER_ASG_PERSON_NUMBER", "PER_ASG_ASSIGNMENT_ID"],
            "contexts": ["DATE_EARNED"],
            "functions": ["TO_CHAR"],
            "return_var": "result_value",
        },
        "rate": {
            "inputs": ["base_amount (NUMBER)", "conversion_rate (NUMBER)"],
            "db_items": ["CMP_ASSIGNMENT_SALARY_AMOUNT"],
            "contexts": ["DATE_EARNED"],
            "functions": ["ROUND"],
            "return_var": "converted_value",
        },
        "payroll": {
            "inputs": ["input_amount (NUMBER)"],
            "db_items": ["PAYROLL_REL_ACTION_ID", "PER_ASG_PERSON_ID"],
            "contexts": ["PAYROLL_REL_ACTION_ID", "DATE_EARNED"],
            "functions": ["ROUND", "GREATEST"],
            "return_var": "result_value",
        },
        "time": {
            "inputs": ["reported_hours (NUMBER)"],
            "db_items": ["HWM_MEASURE_DAY", "HWM_RECORD_POSITION"],
            "contexts": ["DATE_EARNED"],
            "functions": ["ROUND"],
            "return_var": "calculated_hours",
        },
        "generic": {
            "inputs": ["input_value (NUMBER)"],
            "db_items": ["PER_ASG_PERSON_ID"],
            "contexts": ["DATE_EARNED"],
            "functions": ["ROUND"],
            "return_var": "result_value",
        },
    }
    return defaults[kind]


def _select_formula_dbis(preview: str, kind: str) -> List[str]:
    candidates = [
        token
        for token in _extract_formula_database_items(preview.upper())
        if "_" in token and not token.startswith(("FORMULA_", "DEFAULT_", "RETURN_"))
    ]
    selected = list(dict.fromkeys(candidates))[:4]
    if selected:
        return selected
    return list(_default_formula_traits(kind)["db_items"])


def _select_formula_contexts(preview: str, kind: str) -> List[str]:
    contexts = _extract_formula_contexts(preview.upper())
    return contexts or list(_default_formula_traits(kind)["contexts"])


def _select_formula_functions(preview: str, kind: str) -> List[str]:
    functions = _extract_formula_functions(preview.upper())
    return functions or list(_default_formula_traits(kind)["functions"])


def _synthesize_formula_body(
    *,
    formula_type: str,
    use_case: str,
    kind: str,
    db_items: List[str],
    contexts: List[str],
    functions: List[str],
) -> str:
    defaults = _default_formula_traits(kind)
    inputs = defaults["inputs"]
    return_var = defaults["return_var"]
    primary_dbi = db_items[0]
    secondary_dbi = db_items[1] if len(db_items) > 1 else db_items[0]
    primary_fn = functions[0] if functions else ""
    comment = f"/* Derived grounded example for {formula_type}: {use_case} */"

    if kind == "proration":
        body = [
            comment,
            "DEFAULT FOR CALC_START_DATE IS (DATE '1900-01-01')",
            "DEFAULT FOR CALC_END_DATE IS (DATE '4712-12-31')",
            f"INPUTS ARE {', '.join(inputs)}",
            "",
            f"{return_var} = 1",
            "IF start_date > CALC_START_DATE THEN",
            "(",
            "  proration_start = start_date",
            ")",
            "ELSE",
            "(",
            "  proration_start = CALC_START_DATE",
            ")",
            "ENDIF",
            "IF end_date < CALC_END_DATE THEN",
            "(",
            "  proration_end = end_date",
            ")",
            "ELSE",
            "(",
            "  proration_end = CALC_END_DATE",
            ")",
            "ENDIF",
            f"{return_var} = ROUND((DAYS_BETWEEN(proration_end, proration_start) + 1) / 30, 4)",
            f"RETURN {return_var}",
        ]
        return "\n".join(body)

    if kind == "accrual":
        fn = "MONTHS_BETWEEN" if "MONTHS_BETWEEN" in functions or not functions else primary_fn
        body = [
            comment,
            f"DEFAULT FOR {primary_dbi} IS 0",
            "DEFAULT FOR PER_ASG_EFFECTIVE_START_DATE IS (DATE '1900-01-01')",
            f"INPUTS ARE {', '.join(inputs)}",
            "",
            f"{return_var} = 0",
            "IF hire_date WAS DEFAULTED THEN",
            "(",
            "  hire_date = PER_ASG_EFFECTIVE_START_DATE",
            ")",
            "ENDIF",
            f"{return_var} = ROUND({fn}(NVL(term_date, DATE '4712-12-31'), hire_date), 2)",
            f"RETURN {return_var}",
        ]
        return "\n".join(body)

    if kind == "validation":
        body = [
            comment,
            f"DEFAULT FOR {primary_dbi} IS 0",
            f"INPUTS ARE {', '.join(inputs)}",
            "",
            f"{return_var} = 'VALID'",
            "IF entry_value WAS DEFAULTED THEN",
            "(",
            f"  {return_var} = 'INVALID'",
            ")",
            "ENDIF",
            f"RETURN {return_var}",
        ]
        return "\n".join(body)

    if kind == "extract":
        body = [
            comment,
            f"DEFAULT FOR {primary_dbi} IS 'UNKNOWN'",
            f"DEFAULT FOR {secondary_dbi} IS 0",
            f"INPUTS ARE {', '.join(inputs)}",
            "",
            f"{return_var} = TO_CHAR({primary_dbi})",
            f"RETURN {return_var}",
        ]
        return "\n".join(body)

    if kind == "rate":
        body = [
            comment,
            f"DEFAULT FOR {primary_dbi} IS 0",
            f"INPUTS ARE {', '.join(inputs)}",
            "",
            f"{return_var} = ROUND(base_amount * conversion_rate, 2)",
            f"RETURN {return_var}",
        ]
        return "\n".join(body)

    if kind == "time":
        body = [
            comment,
            f"DEFAULT FOR {primary_dbi} IS 0",
            f"INPUTS ARE {', '.join(inputs)}",
            "",
            f"{return_var} = ROUND(reported_hours, 2)",
            f"RETURN {return_var}",
        ]
        return "\n".join(body)

    body = [
        comment,
        f"DEFAULT FOR {primary_dbi} IS 0",
        f"INPUTS ARE {', '.join(inputs)}",
        "",
        f"{return_var} = {primary_dbi}",
        f"RETURN {return_var}",
    ]
    return "\n".join(body)


def _iter_formula_workbook_records(path: Path) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    workbook = load_workbook(path, read_only=True, data_only=True)
    records: List[Dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration:
            continue
        headers = [re.sub(r"\s+", "_", str(value or "").strip()).upper() for value in raw_headers]
        if "FORMULA_TEXT" not in headers or "FORMULA_NAME" not in headers:
            continue

        current: Optional[Dict[str, Any]] = None
        for row in rows:
            item = {header: row[idx] for idx, header in enumerate(headers) if header}
            if item.get("FORMULA_ID") or item.get("FORMULA_NAME"):
                if current and current.get("lines"):
                    current["content"] = "\n".join(current.pop("lines")).strip()
                    records.append(current)
                current = {
                    "formula_id": str(item.get("FORMULA_ID") or ""),
                    "formula_name": str(item.get("FORMULA_NAME") or "").strip(),
                    "formula_type": str(item.get("FORMULA_TYPE_NAME") or "").strip(),
                    "compile_flag": str(item.get("COMPILE_FLAG") or "").strip(),
                    "module_id": str(item.get("MODULE_ID") or "").strip(),
                    "source_sheet": sheet_name,
                    "lines": [],
                }
                continue
            text = item.get("FORMULA_TEXT")
            if current is not None and text is not None:
                current["lines"].append(str(text))

        if current and current.get("lines"):
            current["content"] = "\n".join(current.pop("lines")).strip()
            records.append(current)

    return [record for record in records if record.get("content")]


def _recover_formula_examples_from_workbook(
    path: Path,
    verifier: Verifier,
    known_formula_types: List[str],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    accepted: List[Dict[str, Any]] = []
    rejected: List[Dict[str, Any]] = []
    notes = {
        "workbook_path": str(path),
        "records_detected": 0,
        "real_examples_added": 0,
        "derived_examples_added": 0,
    }
    seen_hashes = set()

    for block in _iter_formula_workbook_records(path):
        notes["records_detected"] += 1
        formula_name = str(block.get("formula_name") or block.get("formula_id") or path.stem).strip()
        formula_type = str(block.get("formula_type") or "").strip()
        formula_type = formula_type if formula_type else _formula_type_from_text(formula_name, known_formula_types)
        formula_type = formula_type if formula_type != "UNKNOWN" else _formula_type_from_text(block["content"], known_formula_types)
        raw_formula = str(block.get("content") or "").strip()
        if not raw_formula:
            rejected.append({"source_path": str(path), "title": formula_name, "reason": "missing_formula_body"})
            continue

        content_hash = stable_hash("formula_workbook_record", formula_name, formula_type, raw_formula)
        if content_hash in seen_hashes:
            continue

        valid, reason = verifier.verify_fast_formula(raw_formula)
        if valid:
            seen_hashes.add(content_hash)
            accepted.append(
                _normalize_formula_record(
                    {
                        "source_path": str(path),
                        "source_uri": f"{path.name}#{block['source_sheet']}#{block.get('formula_id') or formula_name}",
                        "title": formula_name,
                        "module": FusionModule.HCM.value,
                        "task_type": TaskType.FAST_FORMULA_GENERATION.value,
                        "doc_type": DocType.FAST_FORMULA_EXAMPLE,
                        "content": raw_formula,
                        "formula_type": formula_type,
                        "use_case": formula_name,
                    },
                    source_file=path.name,
                    source_sheet=str(block.get("source_sheet") or "workbook"),
                    confidence=0.97,
                )
            )
            notes["real_examples_added"] += 1
            continue

        kind = _formula_template_kind(formula_type, formula_name)
        db_items = _select_formula_dbis(raw_formula, kind)
        contexts = _select_formula_contexts(raw_formula, kind)
        functions = _select_formula_functions(raw_formula, kind)
        synthesized = _synthesize_formula_body(
            formula_type=formula_type,
            use_case=formula_name,
            kind=kind,
            db_items=db_items,
            contexts=contexts,
            functions=functions,
        )
        synth_valid, synth_reason = verifier.verify_fast_formula(synthesized)
        if not synth_valid:
            rejected.append(
                {
                    "source_path": str(path),
                    "title": formula_name,
                    "reason": synth_reason or reason or "formula_verifier_failed",
                }
            )
            continue

        derived_hash = stable_hash("formula_workbook_derived", formula_name, formula_type, synthesized)
        if derived_hash in seen_hashes:
            continue
        seen_hashes.add(derived_hash)
        accepted.append(
            _normalize_formula_record(
                {
                    "source_path": str(path),
                    "source_uri": f"{path.name}#{block['source_sheet']}#{block.get('formula_id') or formula_name}#derived",
                    "title": formula_name,
                    "module": FusionModule.HCM.value,
                    "task_type": TaskType.FAST_FORMULA_GENERATION.value,
                    "doc_type": DocType.FAST_FORMULA_EXAMPLE,
                    "content": synthesized,
                    "formula_type": formula_type,
                    "use_case": formula_name,
                    "derived_from_doc": True,
                    "source_doc": f"{path.name}#{block['source_sheet']}",
                    "confidence_band": "high",
                },
                source_file=path.name,
                source_sheet=str(block.get("source_sheet") or "workbook"),
                confidence=0.9,
            )
        )
        notes["derived_examples_added"] += 1

    return accepted, rejected, notes


def _flatten_section(value: Any, prefix: str = "") -> List[str]:
    lines: List[str] = []
    if isinstance(value, dict):
        for key, child in value.items():
            next_prefix = f"{prefix}{key}: " if prefix else f"{key}: "
            if isinstance(child, (dict, list)):
                lines.extend(_flatten_section(child, next_prefix))
            else:
                lines.append(f"{next_prefix}{child}")
    elif isinstance(value, list):
        for item in value:
            if isinstance(item, (dict, list)):
                lines.extend(_flatten_section(item, prefix))
            else:
                lines.append(f"{prefix}{item}" if prefix else str(item))
    elif value is not None:
        lines.append(f"{prefix}{value}" if prefix else str(value))
    return lines


def _parse_xdm_queries(path: Path) -> List[str]:
    content = _read_text(path)
    queries = re.split(r"\[Query \d+\]", content)
    return [query.strip() for query in queries[1:] if query.strip()]


def _extract_sql_metadata(sql: str, registry: Any) -> Tuple[List[str], List[str], List[str]]:
    try:
        tree = sqlglot.parse_one(sql.strip().rstrip(";"), read="oracle")
    except Exception:
        return [], [], []

    tables = []
    columns = set()
    joins = set()

    seen_tables: List[str] = []
    for table in tree.find_all(exp.Table):
        table_name = registry.resolve_object_name(table.name.upper()) or table.name.upper()
        if table_name not in seen_tables:
            seen_tables.append(table_name)
        tables.append(table_name)

    for column in tree.find_all(exp.Column):
        if column.name:
            columns.add(column.name.upper())

    previous_table = seen_tables[0] if seen_tables else None
    for join in tree.find_all(exp.Join):
        target = join.this
        if not isinstance(target, exp.Table):
            continue
        table_name = registry.resolve_object_name(target.name.upper()) or target.name.upper()
        if previous_table:
            joins.add(f"{previous_table}->{table_name}")
        previous_table = table_name

    return sorted(set(tables)), sorted(columns), sorted(joins)


def _infer_module_from_tables(tables: List[str], registry: Any, source_hint: str) -> str:
    for table_name in tables:
        entry = registry.get_entry(table_name)
        module = str((entry or {}).get("owning_module") or "")
        if module and module not in {"UNKNOWN", "Common"}:
            return module

    for pattern, module in PATH_MODULE_HINTS:
        if pattern.search(source_hint):
            return module
    return FusionModule.COMMON.value


def _infer_sql_task_type(source_hint: str) -> str:
    lowered = source_hint.lower()
    if any(token in lowered for token in ["exception", "exceptions", "error", "variance", "reconciliation", "reco", "failed"]):
        return TaskType.SQL_TROUBLESHOOTING.value
    if any(token in lowered for token in ["report", "analytics", "analysis", "extract", "statement", "dm", "data model"]):
        return TaskType.REPORT_LOGIC.value
    return TaskType.SQL_GENERATION.value


def _normalize_sql_module_label(raw_module: Any) -> str:
    text = str(raw_module or "").strip()
    if not text:
        return FusionModule.COMMON.value
    lowered = text.lower()
    aliases = {
        "accounts payable": FusionModule.PAYABLES.value,
        "payables": FusionModule.PAYABLES.value,
        "accounts receivable": FusionModule.RECEIVABLES.value,
        "receivables": FusionModule.RECEIVABLES.value,
        "general ledger": FusionModule.GENERAL_LEDGER.value,
        "cash management": FusionModule.CASH_MANAGEMENT.value,
        "fixed assets": FusionModule.ASSETS.value,
        "assets": FusionModule.ASSETS.value,
        "expense": FusionModule.EXPENSES.value,
        "expenses": FusionModule.EXPENSES.value,
        "procurement": FusionModule.PROCUREMENT.value,
        "purchasing": FusionModule.PROCUREMENT.value,
        "procure-to-pay": FusionModule.PROCUREMENT.value,
        "order-to-cash": FusionModule.RECEIVABLES.value,
        "inventory": FusionModule.SCM.value,
        "supply chain": FusionModule.SCM.value,
        "scm": FusionModule.SCM.value,
        "hcm": FusionModule.HCM.value,
        "human resources": FusionModule.HCM.value,
        "payroll": FusionModule.HCM.value,
        "projects": FusionModule.PROJECTS.value,
        "project management": FusionModule.PROJECTS.value,
        "tax": FusionModule.TAX.value,
        "financials": FusionModule.COMMON.value,
    }
    if lowered in aliases:
        return aliases[lowered]
    for pattern, module in PATH_MODULE_HINTS:
        if pattern.search(text):
            return module
    return FusionModule.COMMON.value


def _sanitize_oraclewings_sql(sql: str, registry: Any) -> str:
    cleaned = re.sub(r"/\*.*?\*/", " ", sql, flags=re.DOTALL)
    cleaned = re.sub(r"(?m)--[^\n]*$", "", cleaned)
    cleaned = cleaned.replace("\r", "\n")
    for source_name, target_name in ORACLEWINGS_SQL_TABLE_REMAP.items():
        cleaned = re.sub(rf"\b{source_name}\b", target_name, cleaned, flags=re.IGNORECASE)

    aliases = {
        alias.lower()
        for alias in re.findall(r"\b([A-Za-z][A-Za-z0-9_]*)\.", cleaned)
        if len(alias) <= 6
    }

    def _repair_table_token(token: str, suffix: str, prefix: str = "") -> str:
        canonical = registry.resolve_object_name(token.upper()) or token.upper()
        if registry.has_object(canonical):
            return f"{prefix}{canonical}{suffix}"

        token_lower = token.lower()
        for alias in sorted(aliases, key=len, reverse=True):
            if not token_lower.endswith(alias):
                continue
            base_name = token[:-len(alias)]
            if len(base_name) < 3:
                continue
            base_canonical = registry.resolve_object_name(base_name.upper()) or base_name.upper()
            if registry.has_object(base_canonical):
                return f"{prefix}{base_canonical} {alias}{suffix}"
        return f"{prefix}{token}{suffix}"

    def _repair_join_token(match: re.Match[str]) -> str:
        keyword = match.group(1)
        token = match.group(2)
        suffix = match.group(3)
        return _repair_table_token(token, suffix, prefix=f"{keyword} ")

    def _repair_comma_token(match: re.Match[str]) -> str:
        token = match.group(1)
        suffix = match.group(2)
        return _repair_table_token(token, suffix, prefix=", ")

    cleaned = re.sub(r"(?i)\b(from|join)\s+([A-Za-z][A-Za-z0-9_$]*)(\s+)", _repair_join_token, cleaned)
    cleaned = re.sub(r"(?i),\s*([A-Za-z][A-Za-z0-9_$]*)(\s*(?:,|\s))", _repair_comma_token, cleaned)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned).strip()
    return cleaned


def _repair_oraclewings_sql_with_verifier_reason(sql: str, reason: str) -> str:
    reason_text = str(reason or "")
    repaired = sql
    rule_fired = False

    def _replace(pattern: str, replacement: str) -> None:
        nonlocal repaired, rule_fired
        next_sql = re.sub(pattern, replacement, repaired, flags=re.IGNORECASE)
        if next_sql != repaired:
            repaired = next_sql
            rule_fired = True

    if "Column 'AUTHORIZATION_STATUS' is not present on table 'PO_HEADERS_ALL'" in reason_text:
        _replace(r"\bAUTHORIZATION_STATUS\b", "DOCUMENT_STATUS")

    if "Column 'SEGMENT1' is not present on table 'POR_REQUISITION_HEADERS_ALL'" in reason_text:
        _replace(r"\bSEGMENT1\b", "REQUISITION_NUMBER")

    if "Column 'EMPLOYEE_NUMBER' is not present on table 'PER_ALL_PEOPLE_F'" in reason_text:
        _replace(r"\bEMPLOYEE_NUMBER\b", "PERSON_NUMBER")

    if "Column 'FULL_NAME' is not present on table 'PER_ALL_PEOPLE_F'" in reason_text:
        _replace(r"\bFULL_NAME\b", "PERSON_NUMBER")

    if "Column 'NAME' is not present on table 'FUN_ALL_BUSINESS_UNITS_V'" in reason_text:
        _replace(r"\bNAME\b", "BU_NAME")

    if "Column 'VENDOR_NAME' is not present on table 'POZ_SUPPLIERS'" in reason_text:
        _replace(r"\bPOZ_SUPPLIERS\b", "POZ_SUPPLIERS_V")

    if "Table 'FA_CATEGORIES' is not present in the Oracle Fusion metadata index." in reason_text:
        _replace(r"\bFA_CATEGORIES\b", "FA_CATEGORIES_B")

    return repaired if rule_fired else sql


def _inventory_sql_examples(registry: Any, verifier: Verifier) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    accepted: List[Dict[str, Any]] = []
    rejected: List[Dict[str, Any]] = []
    seen_hashes = set()

    for path in sorted(ORG_SQL_DIR.rglob("*.xdm.txt")):
        rel_path = str(path.relative_to(ORG_SQL_DIR))
        queries = _parse_xdm_queries(path)
        if not queries:
            rejected.append({"source_path": str(path), "reason": "empty_or_unparsed_sql"})
            continue

        for idx, sql in enumerate(queries, start=1):
            content_hash = stable_hash("sql_examples_corpus", sql)
            title = f"{_clean_title(path.stem)} [Q{idx}]"

            reject_reason = CuratedIngestionValidator.reject_sql(sql)
            if reject_reason:
                rejected.append({"source_path": str(path), "title": title, "reason": reject_reason})
                continue

            valid, reason = verifier.verify_sql(sql)
            if not valid:
                rejected.append({"source_path": str(path), "title": title, "reason": reason or "sql_verifier_failed"})
                continue

            tables_used, columns_used, joins_used = _extract_sql_metadata(sql, registry)
            if not tables_used:
                rejected.append({"source_path": str(path), "title": title, "reason": "no_grounded_tables"})
                continue

            module = _infer_module_from_tables(tables_used, registry, rel_path)
            if module in {"UNKNOWN", FusionModule.COMMON.value}:
                rejected.append({"source_path": str(path), "title": title, "reason": "unclassifiable_module"})
                continue

            if content_hash in seen_hashes:
                rejected.append({"source_path": str(path), "title": title, "reason": "duplicate_sql"})
                continue
            seen_hashes.add(content_hash)

            confidence = 0.98 if any(
                str((registry.get_entry(table_name) or {}).get("owning_module") or "") == module
                for table_name in tables_used
            ) else 0.88

            accepted.append(
                {
                    "source_path": str(path),
                    "source_uri": rel_path,
                    "title": title,
                    "module": module,
                    "task_type": _infer_sql_task_type(rel_path),
                    "content": sql,
                    "tables_used": tables_used,
                    "columns_used": columns_used,
                    "joins_used": joins_used,
                    "source_file": rel_path,
                    "confidence": round(confidence, 2),
                    "content_hash": content_hash,
                }
            )

    return accepted, rejected


def _inventory_oraclewings_sql_examples(registry: Any, verifier: Verifier) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    accepted: List[Dict[str, Any]] = []
    rejected: List[Dict[str, Any]] = []
    seen_hashes = set()

    if not ORACLEWINGS_FUSION_QUERIES_PATH.exists():
        return accepted, rejected

    payload = json.loads(_read_text(ORACLEWINGS_FUSION_QUERIES_PATH) or "{}")
    rows = payload.get("queries") if isinstance(payload, dict) else []
    if not isinstance(rows, list):
        return accepted, rejected

    source_rel = str(ORACLEWINGS_FUSION_QUERIES_PATH.relative_to(ROOT_DIR))
    for idx, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            continue
        sql = _sanitize_oraclewings_sql(str(row.get("sql") or "").strip(), registry)
        if not sql:
            rejected.append({"source_path": str(ORACLEWINGS_FUSION_QUERIES_PATH), "title": f"Fusion Query {idx}", "reason": "missing_sql"})
            continue

        title = _clean_title(str(row.get("description") or row.get("intent") or f"Fusion Query {idx}"))
        reject_reason = CuratedIngestionValidator.reject_sql(sql)
        if reject_reason:
            rejected.append({"source_path": str(ORACLEWINGS_FUSION_QUERIES_PATH), "title": title, "reason": reject_reason})
            continue

        valid, reason = verifier.verify_sql(sql)
        if not valid:
            repaired_sql = _repair_oraclewings_sql_with_verifier_reason(sql, reason or "")
            if repaired_sql != sql:
                valid, reason = verifier.verify_sql(repaired_sql)
                if valid:
                    sql = repaired_sql
            if not valid:
                rejected.append({"source_path": str(ORACLEWINGS_FUSION_QUERIES_PATH), "title": title, "reason": reason or "sql_verifier_failed"})
                continue

        tables_used, columns_used, joins_used = _extract_sql_metadata(sql, registry)
        if not tables_used:
            hinted_tables = []
            for table_name in row.get("tables") or []:
                canonical = registry.resolve_object_name(str(table_name).upper()) or str(table_name).upper()
                if registry.get_entry(canonical):
                    hinted_tables.append(canonical)
            tables_used = sorted(set(hinted_tables))

        if not tables_used:
            rejected.append({"source_path": str(ORACLEWINGS_FUSION_QUERIES_PATH), "title": title, "reason": "no_grounded_tables"})
            continue

        module = _normalize_sql_module_label(row.get("module"))
        if module in {"UNKNOWN", FusionModule.COMMON.value}:
            module = _infer_module_from_tables(tables_used, registry, f"{row.get('module', '')} {title}")
        if module in {"UNKNOWN", FusionModule.COMMON.value}:
            rejected.append({"source_path": str(ORACLEWINGS_FUSION_QUERIES_PATH), "title": title, "reason": "unclassifiable_module"})
            continue

        if not joins_used and len(tables_used) > 1:
            joins_used = [f"{tables_used[pos]}->{tables_used[pos + 1]}" for pos in range(len(tables_used) - 1)]

        content_hash = stable_hash("oraclewings_fusion_queries", sql)
        if content_hash in seen_hashes:
            rejected.append({"source_path": str(ORACLEWINGS_FUSION_QUERIES_PATH), "title": title, "reason": "duplicate_sql"})
            continue
        seen_hashes.add(content_hash)

        accepted.append(
            {
                "source_path": str(ORACLEWINGS_FUSION_QUERIES_PATH),
                "source_uri": f"{source_rel}#Q{idx}",
                "title": f"{title} [Oraclewings Q{idx}]",
                "module": module,
                "task_type": _infer_sql_task_type(f"{row.get('intent') or ''} {title}"),
                "content": sql,
                "tables_used": tables_used,
                "columns_used": columns_used,
                "joins_used": joins_used,
                "source_file": source_rel,
                "confidence": 0.94,
                "content_hash": content_hash,
                "derived_from_doc": True,
                "source_doc": source_rel,
            }
        )

    return accepted, rejected


def _build_sql_example_chunks(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    chunks: List[Dict[str, Any]] = []
    for record in records:
        document = CuratedIngestionValidator.build_document(
            source_path=record["source_path"],
            source_uri=record["source_uri"],
            title=record["title"],
            module=record["module"],
            task_type=record["task_type"],
            doc_type=DocType.SQL_EXAMPLE,
            trusted_schema_objects=list(record["tables_used"]),
            quality_score=float(record["confidence"]),
            source_system=SourceSystem.ORACLEWINGS_REPO,
            content=record["content"],
            document_id=f"SQLX::{record['content_hash'][:16]}",
            metadata={
                "source_file": record["source_file"],
                "tables_used": list(record["tables_used"]),
                "columns_used": list(record["columns_used"]),
                "joins_used": list(record["joins_used"]),
                "confidence": float(record["confidence"]),
                "derived_from_doc": bool(record.get("derived_from_doc")),
                "derived_from_runtime": bool(record.get("derived_from_runtime")),
                "source_doc": str(record.get("source_doc") or ""),
                "authority_tier": "secondary",
            },
        )
        chunk = CuratedIngestionValidator.build_chunk(document, record["content"], 0)
        chunks.append(CuratedIngestionValidator.chunk_payload(chunk))
    return chunks


def _build_schema_metadata_chunks(registry: Any, sql_records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    tables_of_interest = set()
    for record in sql_records:
        tables_of_interest.update(record["tables_used"])
        for table_name in record["tables_used"]:
            tables_of_interest.update(registry.get_related_objects(table_name))

    chunks: List[Dict[str, Any]] = []
    relation_pairs = set()
    for table_name in sorted(tables_of_interest):
        entry = registry.get_entry(table_name)
        if not entry:
            continue
        module = str(entry.get("owning_module") or entry.get("inferred_module") or FusionModule.COMMON.value)
        aliases = sorted(set(entry.get("aliases", []) + entry.get("ebs_aliases", [])))[:12]
        relations = sorted(set(entry.get("approved_relations", [])))[:20]
        primary_keys = registry.get_primary_keys(table_name)[:12]
        base_tables = registry.get_view_base_tables(table_name)[:12]
        lines = [
            f"OBJECT: {table_name}",
            f"OBJECT_TYPE: {entry.get('object_type')}",
            f"MODULE: {module}",
            f"MODULE_FAMILY: {entry.get('owning_module_family')}",
        ]
        if aliases:
            lines.append("ALIASES: " + ", ".join(aliases))
        if entry.get("columns"):
            lines.append("COLUMNS: " + ", ".join(entry["columns"][:50]))
        if primary_keys:
            lines.append("PRIMARY_KEYS: " + ", ".join(primary_keys))
        if base_tables:
            lines.append("BASE_TABLES: " + ", ".join(base_tables))
        if relations:
            lines.append("RELATED_OBJECTS: " + ", ".join(relations))

        document = CuratedIngestionValidator.build_document(
            source_path="trusted_registry",
            source_uri=f"registry://{table_name}",
            title=table_name,
            module=module,
            task_type=TaskType.TABLE_LOOKUP.value,
            doc_type=DocType.SCHEMA_METADATA,
            trusted_schema_objects=[table_name, *relations[:10]],
            quality_score=0.97,
            source_system=SourceSystem.METADATA,
            content="\n".join(lines),
            document_id=f"SCHEMAX::{table_name}",
            metadata={
                "aliases": aliases,
                "columns_used": entry.get("columns", [])[:50],
                "primary_keys": primary_keys,
                "base_tables": base_tables,
                "related_objects": relations,
                "authority_tier": "official",
            },
        )
        chunk = CuratedIngestionValidator.build_chunk(document, "\n".join(lines), 0)
        chunks.append(CuratedIngestionValidator.chunk_payload(chunk))

        for related in relations:
            pair = tuple(sorted((table_name, related)))
            if pair in relation_pairs:
                continue
            relation_pairs.add(pair)
            related_entry = registry.get_entry(related)
            related_module = str((related_entry or {}).get("owning_module") or FusionModule.COMMON.value)
            relation_details = registry.get_relation_details(table_name, related)
            join_lines = []
            for detail in relation_details[:6]:
                src_table = str(detail.get("source_table") or "").upper()
                src_column = str(detail.get("source_column") or "").upper()
                tgt_table = str(detail.get("target_table") or "").upper()
                tgt_column = str(detail.get("target_column") or "").upper()
                if src_table and src_column and tgt_table and tgt_column:
                    join_lines.append(f"JOIN: {src_table}.{src_column} = {tgt_table}.{tgt_column}")
            relation_content = (
                f"RELATION: {pair[0]} <-> {pair[1]}\n"
                f"PRIMARY_OBJECT_MODULE: {module}\n"
                f"RELATED_OBJECT_MODULE: {related_module}\n"
                "RELATION_SOURCE: trusted Oracle Fusion metadata foreign-key mapping"
            )
            if join_lines:
                relation_content += "\n" + "\n".join(join_lines)
            relation_document = CuratedIngestionValidator.build_document(
                source_path="trusted_registry",
                source_uri=f"registry://relation/{pair[0]}::{pair[1]}",
                title=f"{pair[0]} to {pair[1]} relation",
                module=module if module != FusionModule.COMMON.value else related_module,
                task_type=TaskType.TABLE_LOOKUP.value,
                doc_type=DocType.SCHEMA_METADATA,
                trusted_schema_objects=[pair[0], pair[1]],
                quality_score=0.95,
                source_system=SourceSystem.METADATA,
                content=relation_content,
                document_id=f"SCHEMAXREL::{pair[0]}::{pair[1]}",
                metadata={
                    "authority_tier": "official",
                    "relation_details": relation_details[:10],
                },
            )
            relation_chunk = CuratedIngestionValidator.build_chunk(relation_document, relation_content, 0)
            chunks.append(CuratedIngestionValidator.chunk_payload(relation_chunk))

    return chunks


def _parse_formula_csv_blocks(path: Path) -> List[Dict[str, Any]]:
    text = _read_text(path)
    blocks: List[Dict[str, Any]] = []
    current_title: Optional[str] = None
    current_lines: List[str] = []
    start_pattern = re.compile(r"^(?:\d+\.\d+\s+Template\s+—\s+.+FORMULA.*|✅\s*Example\s+\d+\s+—\s+.+)$", re.IGNORECASE)
    section_break_pattern = re.compile(r"^\d+\.\s+[A-Z].+$")

    def flush() -> None:
        nonlocal current_title, current_lines
        if not current_title:
            current_lines = []
            return
        content = "\n".join(line.rstrip() for line in current_lines).strip()
        normalized = "\n".join(
            line for line in (_normalize_formula_line(line) for line in content.splitlines()) if line
        ).strip()
        if "RETURN" in normalized.upper() and any(
            token in normalized.upper() for token in ("FORMULA NAME:", "INPUTS ARE", "DEFAULT FOR", "IF ", "WHILE ")
        ):
            blocks.append({"title": current_title, "content": normalized})
        current_title = None
        current_lines = []

    for raw_line in text.splitlines():
        line = raw_line.rstrip()
        stripped = line.strip()
        if start_pattern.match(stripped):
            flush()
            current_title = stripped
            continue

        if current_title and section_break_pattern.match(stripped) and not start_pattern.match(stripped):
            flush()
            continue

        if current_title is not None:
            current_lines.append(line)

    flush()
    return blocks


def _load_formula_types() -> List[str]:
    if not FORMULA_TYPES_PATH.exists():
        return []
    payload = json.loads(_read_text(FORMULA_TYPES_PATH))
    if isinstance(payload, list):
        values: List[str] = []
        for item in payload:
            if isinstance(item, dict):
                name = str(item.get("name") or item.get("formula_type") or "").strip()
                if name:
                    values.append(name)
            else:
                text = str(item).strip()
                if text:
                    values.append(text)
        return values
    return []


def _derive_formula_examples_from_harvest(
    verifier: Verifier,
    known_formula_types: List[str],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    accepted: List[Dict[str, Any]] = []
    rejected: List[Dict[str, Any]] = []
    notes = {
        "support_docs_considered": 0,
        "derived_examples_created": 0,
        "source_docs_used": 0,
    }
    rows = _read_jsonl(HARVEST_FORMULA_SUPPORT_PATH)
    if not rows:
        return accepted, rejected, notes

    seen_titles = set()
    source_docs = set()
    for row in rows:
        payload = row.get("normalized_payload") or {}
        preview = _harvest_preview_text(row)
        formula_type_raw = payload.get("formula_type") or row.get("title") or "UNKNOWN"
        formula_type, description = _normalize_harvest_formula_type(formula_type_raw)
        formula_type = formula_type if formula_type != "UNKNOWN" else _formula_type_from_text(preview, known_formula_types)
        use_case = description or str(payload.get("section_title") or row.get("title") or formula_type)
        kind = _formula_template_kind(formula_type, use_case)
        db_items = _select_formula_dbis(preview, kind)
        contexts = _select_formula_contexts(preview, kind)
        functions = _select_formula_functions(preview, kind)
        formula_body = _synthesize_formula_body(
            formula_type=formula_type,
            use_case=use_case,
            kind=kind,
            db_items=db_items,
            contexts=contexts,
            functions=functions,
        )

        valid, reason = verifier.verify_fast_formula(formula_body)
        notes["support_docs_considered"] += 1
        if not valid:
            rejected.append(
                {
                    "source_path": row.get("source_path"),
                    "title": row.get("title"),
                    "reason": reason or "derived_formula_verifier_failed",
                }
            )
            continue

        title = f"{formula_type} Derived Example"
        if title in seen_titles:
            title = f"{formula_type} Derived Example ({_clean_title(use_case)})"
        if title in seen_titles:
            continue
        seen_titles.add(title)
        source_doc = f"{row.get('source_path')}#{row.get('title')}"
        source_docs.add(source_doc)
        accepted.append(
            _normalize_formula_record(
                {
                    "source_path": str(row.get("source_path") or HARVEST_FORMULA_SUPPORT_PATH),
                    "source_uri": source_doc,
                    "title": title,
                    "module": FusionModule.HCM.value,
                    "task_type": TaskType.FAST_FORMULA_GENERATION.value,
                    "doc_type": DocType.FAST_FORMULA_EXAMPLE,
                    "content": formula_body,
                    "formula_type": formula_type,
                    "use_case": _clean_title(use_case),
                    "derived_from_doc": True,
                    "source_doc": source_doc,
                    "confidence_band": "high" if payload.get("source_kind") == "json_support_section" else "medium",
                },
                source_file=Path(str(row.get("source_path") or "harvest")).name,
                source_sheet=str(payload.get("section_title") or payload.get("source_kind") or "harvest_doc"),
                confidence=0.86 if payload.get("source_kind") == "json_support_section" else 0.82,
            )
        )

    notes["derived_examples_created"] = len(accepted)
    notes["source_docs_used"] = len(source_docs)
    return accepted, rejected, notes


def _load_harvest_formula_examples(known_formula_types: List[str]) -> List[Dict[str, Any]]:
    accepted: List[Dict[str, Any]] = []
    for row in _read_jsonl(HARVEST_FORMULA_EXAMPLES_PATH):
        payload = row.get("normalized_payload") or {}
        formula_body = str(payload.get("formula_body") or "").strip()
        if not formula_body:
            continue
        formula_type = str(payload.get("formula_type") or _formula_type_from_text(_harvest_preview_text(row), known_formula_types))
        accepted.append(
            _normalize_formula_record(
                {
                    "source_path": str(row.get("source_path") or HARVEST_FORMULA_EXAMPLES_PATH),
                    "source_uri": str(row.get("source_path") or HARVEST_FORMULA_EXAMPLES_PATH),
                    "title": str(row.get("title") or payload.get("formula_name") or "Harvest Formula Example"),
                    "module": str(row.get("module") or FusionModule.HCM.value),
                    "task_type": TaskType.FAST_FORMULA_GENERATION.value,
                    "doc_type": DocType.FAST_FORMULA_EXAMPLE,
                    "content": formula_body,
                    "formula_type": formula_type,
                    "use_case": str(payload.get("use_case") or row.get("title") or formula_type),
                    "source_doc": str(row.get("source_path") or ""),
                },
                source_file=Path(str(row.get("source_path") or "harvest")).name,
                source_sheet=str(payload.get("source_sheet") or payload.get("source_kind") or "harvest_examples"),
                confidence=float(row.get("confidence") or 0.9),
            )
        )
    return accepted


def _preferred_registry_columns(entry: Dict[str, Any], limit: int = 6) -> List[str]:
    columns = [str(column).upper() for column in entry.get("columns", [])]
    preferred = [
        column
        for column in columns
        if not column.startswith(AUDIT_COLUMN_PREFIXES)
        and not re.match(r"ATTRIBUTE\d+$", column)
    ]
    selected = preferred[:limit]
    if len(selected) < min(limit, len(columns)):
        for column in columns:
            if column not in selected:
                selected.append(column)
            if len(selected) >= limit:
                break
    return selected


def _build_single_table_sql(table_name: str, columns: List[str], *, troubleshooting: bool) -> str:
    projected = columns[:5] if columns else ["*"]
    if not projected:
        projected = ["*"]
    where_clauses = []
    status_cols = [column for column in projected if "STATUS" in column]
    if troubleshooting and status_cols:
        where_clauses.append(f"{status_cols[0]} IS NOT NULL")
    where_sql = f"\nWHERE {' AND '.join(where_clauses)}" if where_clauses else ""
    return f"SELECT {', '.join(projected)}\nFROM {table_name}{where_sql}\nFETCH FIRST 100 ROWS ONLY"


def _pick_join_relation(registry: Any, left_table: str, right_table: str) -> Optional[Dict[str, Any]]:
    details = registry.get_relation_details(left_table, right_table)
    if not details:
        return None
    valid = [
        detail
        for detail in details
        if str(detail.get("source_column") or "").strip()
        and str(detail.get("target_column") or "").strip()
    ]
    if not valid:
        return None
    valid.sort(
        key=lambda item: (
            0 if str(item.get("fk_name") or "").upper().endswith("_PK") else 1,
            len(str(item.get("source_column") or "")),
            str(item.get("fk_name") or ""),
        )
    )
    return valid[0]


def _build_join_sql(
    registry: Any,
    left_table: str,
    right_table: str,
    *,
    troubleshooting: bool,
) -> Optional[str]:
    relation = _pick_join_relation(registry, left_table, right_table)
    if not relation:
        return None

    source_table = str(relation.get("source_table") or "").upper()
    source_column = str(relation.get("source_column") or "").upper()
    target_table = str(relation.get("target_table") or "").upper()
    target_column = str(relation.get("target_column") or "").upper()
    if not all([source_table, source_column, target_table, target_column]):
        return None

    left_entry = registry.get_entry(left_table)
    right_entry = registry.get_entry(right_table)
    if not left_entry or not right_entry:
        return None

    left_columns = _preferred_registry_columns(left_entry, limit=4)
    right_columns = _preferred_registry_columns(right_entry, limit=4)
    if not left_columns or not right_columns:
        return None

    projections = [f"a.{column}" for column in left_columns[:3]] + [f"b.{column}" for column in right_columns[:3]]
    join_left_column, join_right_column = (
        (source_column, target_column)
        if left_table == source_table and right_table == target_table
        else (target_column, source_column)
    )
    where_clauses = []
    if troubleshooting:
        status_candidates = [
            f"a.{column}"
            for column in left_columns
            if "STATUS" in column or "ERROR" in column
        ] + [
            f"b.{column}"
            for column in right_columns
            if "STATUS" in column or "ERROR" in column
        ]
        if status_candidates:
            where_clauses.append(f"{status_candidates[0]} IS NOT NULL")
    where_sql = f"\nWHERE {' AND '.join(where_clauses)}" if where_clauses else ""
    return (
        f"SELECT {', '.join(projections)}\n"
        f"FROM {left_table} a\n"
        f"JOIN {right_table} b\n"
        f"  ON a.{join_left_column} = b.{join_right_column}"
        f"{where_sql}\n"
        "FETCH FIRST 100 ROWS ONLY"
    )


def _derive_join_graph_sql_patterns(
    registry: Any,
    verifier: Verifier,
    seed_records: List[Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    accepted: List[Dict[str, Any]] = []
    rejected: List[Dict[str, Any]] = []
    seen_hashes = set()
    candidate_pairs = set()
    candidate_counts: Counter[Tuple[str, str]] = Counter()

    for record in seed_records:
        tables = [table for table in record.get("tables_used", []) if registry.resolve_object_name(table)]
        if len(tables) >= 2:
            for idx in range(len(tables) - 1):
                pair = tuple(sorted((tables[idx], tables[idx + 1])))
                candidate_pairs.add(pair)
                candidate_counts[pair] += 2
        for table_name in tables[:4]:
            related = registry.get_related_objects(table_name)
            for related_name in related[:4]:
                pair = tuple(sorted((table_name, related_name)))
                candidate_pairs.add(pair)
                candidate_counts[pair] += 1

    for left_table, right_table in sorted(candidate_pairs, key=lambda pair: (-candidate_counts[pair], pair)):
        join_path = registry.find_join_path(left_table, right_table, max_depth=2)
        if join_path != [left_table, right_table] and join_path != [right_table, left_table]:
            continue
        sql = _build_join_sql(
            registry,
            left_table,
            right_table,
            troubleshooting=False,
        )
        if not sql:
            rejected.append(
                {
                    "source_path": "registry://join_graph",
                    "title": f"{left_table} to {right_table}",
                    "reason": "missing_relation_detail",
                }
            )
            continue
        valid, reason = verifier.verify_sql(sql)
        if not valid:
            rejected.append(
                {
                    "source_path": "registry://join_graph",
                    "title": f"{left_table} to {right_table}",
                    "reason": reason or "sql_verifier_failed",
                }
            )
            continue
        tables_used, columns_used, joins_used = _extract_sql_metadata(sql, registry)
        module = _infer_module_from_tables(
            tables_used,
            registry,
            f"{left_table} {right_table} relation",
        )
        if module in {"UNKNOWN", FusionModule.COMMON.value}:
            rejected.append(
                {
                    "source_path": "registry://join_graph",
                    "title": f"{left_table} to {right_table}",
                    "reason": "unclassifiable_module",
                }
            )
            continue
        content_hash = stable_hash("join_graph_sql", left_table, right_table, sql)
        if content_hash in seen_hashes:
            continue
        seen_hashes.add(content_hash)
        accepted.append(
            {
                "source_path": "trusted_registry",
                "source_uri": f"registry://relation/{left_table}::{right_table}",
                "title": f"{left_table} to {right_table} join template",
                "module": module,
                "task_type": TaskType.SQL_GENERATION.value,
                "content": sql,
                "tables_used": tables_used,
                "columns_used": columns_used,
                "joins_used": joins_used,
                "source_file": f"{left_table}::{right_table}",
                "confidence": 0.87,
                "content_hash": content_hash,
                "derived_from_doc": True,
                "source_doc": f"registry://relation/{left_table}::{right_table}",
            }
        )

    return accepted, rejected, {"join_graph_patterns_added": len(accepted), "candidate_pairs": len(candidate_pairs)}


def _derive_sql_patterns_from_harvest(
    registry: Any,
    verifier: Verifier,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    accepted: List[Dict[str, Any]] = []
    rejected: List[Dict[str, Any]] = []
    notes = {
        "hdl_rows_considered": 0,
        "runtime_rows_considered": 0,
        "patterns_from_hdl": 0,
        "patterns_from_runtime": 0,
    }
    seen_hashes = set()

    def maybe_add_record(
        *,
        source_row: Dict[str, Any],
        title: str,
        table_name: str,
        source_uri: str,
        task_type: str,
        derived_from_doc: bool = False,
        derived_from_runtime: bool = False,
        confidence: float,
    ) -> None:
        sanitized_table_name = _sanitize_harvest_table_name(table_name)
        canonical_table = registry.resolve_object_name(sanitized_table_name or table_name)
        if not canonical_table:
            rejected.append({"source_path": source_row.get("source_path"), "title": title, "reason": "unverified_table"})
            return
        entry = registry.get_entry(canonical_table)
        if not entry:
            rejected.append({"source_path": source_row.get("source_path"), "title": title, "reason": "missing_registry_entry"})
            return
        module = str(entry.get("owning_module") or entry.get("inferred_module") or FusionModule.COMMON.value)
        if module in {FusionModule.UNKNOWN.value, FusionModule.COMMON.value, "UNKNOWN", "Common"}:
            source_hint = " ".join(
                [
                    str(source_row.get("title") or ""),
                    str(source_row.get("module") or ""),
                    str(source_row.get("subsystem") or ""),
                    str(source_row.get("source_path") or ""),
                ]
            )
            module = _infer_module_from_tables([canonical_table], registry, source_hint)
            if module in {FusionModule.UNKNOWN.value, FusionModule.COMMON.value, "UNKNOWN", "Common"}:
                rejected.append({"source_path": source_row.get("source_path"), "title": title, "reason": "unclassifiable_module"})
                return
        columns = _preferred_registry_columns(entry)
        if not columns:
            rejected.append({"source_path": source_row.get("source_path"), "title": title, "reason": "no_grounded_columns"})
            return
        sql = _build_single_table_sql(canonical_table, columns, troubleshooting=task_type == TaskType.SQL_TROUBLESHOOTING.value)
        valid, reason = verifier.verify_sql(sql)
        if not valid:
            rejected.append({"source_path": source_row.get("source_path"), "title": title, "reason": reason or "sql_verifier_failed"})
            return
        tables_used, columns_used, joins_used = _extract_sql_metadata(sql, registry)
        content_hash = stable_hash("derived_sql_example", title, sql)
        if content_hash in seen_hashes:
            rejected.append({"source_path": source_row.get("source_path"), "title": title, "reason": "duplicate_sql"})
            return
        seen_hashes.add(content_hash)
        accepted.append(
            {
                "source_path": str(source_row.get("source_path") or HARVEST_HDL_FBDI_PATH),
                "source_uri": source_uri,
                "title": title,
                "module": module,
                "task_type": task_type,
                "content": sql,
                "tables_used": tables_used,
                "columns_used": columns_used,
                "joins_used": joins_used,
                "source_file": str(source_row.get("source_path") or ""),
                "confidence": round(confidence, 2),
                "content_hash": content_hash,
                "derived_from_doc": derived_from_doc,
                "derived_from_runtime": derived_from_runtime,
                "source_doc": source_uri,
            }
        )

    for row in _read_jsonl(HARVEST_HDL_FBDI_PATH):
        payload = row.get("normalized_payload") or {}
        table_name = _sanitize_harvest_table_name(payload.get("table_name") or "")
        if not table_name:
            continue
        notes["hdl_rows_considered"] += 1
        title = f"{row.get('title')} - {table_name}"
        task_type = _infer_sql_task_type(f"{row.get('title')} {row.get('source_path')}")
        maybe_add_record(
            source_row=row,
            title=title,
            table_name=table_name,
            source_uri=f"harvest://hdl_fbdi/{row.get('source_path')}#{table_name}",
            task_type=task_type,
            derived_from_doc=True,
            confidence=0.84,
        )

    notes["patterns_from_hdl"] = len(accepted)

    before_runtime = len(accepted)
    for row in _read_jsonl(HARVEST_SQL_RUNTIME_PATH):
        payload = row.get("normalized_payload") or {}
        summary_text = str(payload.get("extractable_text_summary") or "")
        asset_type = str(payload.get("asset_type") or "")
        tables = []
        for token in sorted(set(re.findall(r"\b[A-Z][A-Z0-9_]{3,}\b", summary_text))):
            if registry.has_object(token):
                tables.append(token)
        if not tables:
            continue
        notes["runtime_rows_considered"] += 1
        task_type = (
            TaskType.SQL_TROUBLESHOOTING.value
            if asset_type in {"validator", "runtime_asset"}
            else TaskType.SQL_GENERATION.value
        )
        for table_name in tables[:6]:
            title = f"{row.get('title')} reference - {table_name}"
            maybe_add_record(
                source_row=row,
                title=title,
                table_name=table_name,
                source_uri=f"harvest://sql_runtime/{row.get('source_path')}#{table_name}",
                task_type=task_type,
                derived_from_runtime=True,
                confidence=0.8,
            )

    notes["patterns_from_runtime"] = len(accepted) - before_runtime
    return accepted, rejected, notes


def _dedupe_sql_records(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    deduped: List[Dict[str, Any]] = []
    seen = set()
    for record in records:
        content_hash = str(record.get("content_hash") or stable_hash("sql_records", record.get("content") or ""))
        if content_hash in seen:
            continue
        seen.add(content_hash)
        deduped.append(record)
    return deduped


def _dedupe_formula_records(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    deduped: List[Dict[str, Any]] = []
    seen = set()
    for record in records:
        content_hash = stable_hash("formula_records", record.get("content") or "")
        if content_hash in seen:
            continue
        seen.add(content_hash)
        deduped.append(record)
    return deduped


def _inventory_fast_formula_assets(verifier: Verifier) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
    accepted: List[Dict[str, Any]] = []
    rejected: List[Dict[str, Any]] = []
    notes: Dict[str, Any] = {
        "candidate_assets_found": [],
        "classified_assets": {
            "fast_formula_examples": [],
            "fast_formula_supporting_docs": [],
            "reject_corpus": [],
        },
    }
    known_formula_types = _load_formula_types()

    if FORMULA_KNOWLEDGE_PATH.exists():
        notes["candidate_assets_found"].append(str(FORMULA_KNOWLEDGE_PATH))
        payload = json.loads(_read_text(FORMULA_KNOWLEDGE_PATH))
        examples = payload.get("formula_examples") or {}
        for key, value in examples.items():
            if not isinstance(value, dict):
                rejected.append({"source_path": str(FORMULA_KNOWLEDGE_PATH), "title": str(key), "reason": "malformed_formula_example"})
                notes["classified_assets"]["reject_corpus"].append(str(FORMULA_KNOWLEDGE_PATH))
                continue
            formula = str(value.get("formula") or "").strip()
            if not formula:
                rejected.append({"source_path": str(FORMULA_KNOWLEDGE_PATH), "title": str(key), "reason": "missing_formula_body"})
                notes["classified_assets"]["reject_corpus"].append(str(FORMULA_KNOWLEDGE_PATH))
                continue
            valid, reason = verifier.verify_fast_formula(formula)
            if not valid:
                rejected.append({"source_path": str(FORMULA_KNOWLEDGE_PATH), "title": str(key), "reason": reason or "formula_verifier_failed"})
                notes["classified_assets"]["reject_corpus"].append(str(FORMULA_KNOWLEDGE_PATH))
                continue
            description = str(value.get("description") or key)
            accepted.append(
                _normalize_formula_record(
                    {
                        "source_path": str(FORMULA_KNOWLEDGE_PATH),
                        "source_uri": "FAST_FORMULA_KNOWLEDGE.json",
                        "title": _clean_title(key),
                        "module": FusionModule.HCM.value,
                        "task_type": TaskType.FAST_FORMULA_GENERATION.value,
                        "doc_type": DocType.FAST_FORMULA_EXAMPLE,
                        "content": formula,
                        "formula_type": _formula_type_from_text(f"{key} {description} {formula}", known_formula_types),
                        "use_case": description,
                    },
                    source_file=FORMULA_KNOWLEDGE_PATH.name,
                    source_sheet="formula_examples",
                    confidence=0.92,
                )
            )
            notes["classified_assets"]["fast_formula_examples"].append(str(FORMULA_KNOWLEDGE_PATH))

        support_sections = [
            ("formula_types_and_contexts", TaskType.FAST_FORMULA_GENERATION.value),
            ("database_items", TaskType.FAST_FORMULA_GENERATION.value),
            ("functions", TaskType.FAST_FORMULA_GENERATION.value),
            ("debugging_and_troubleshooting", TaskType.FAST_FORMULA_TROUBLESHOOTING.value),
        ]
        for key, task_type in support_sections:
            section = payload.get(key)
            if not section:
                continue
            section_lines = _flatten_section(section)
            if not section_lines:
                continue
            accepted.append(
                _normalize_formula_record(
                    {
                        "source_path": str(FORMULA_KNOWLEDGE_PATH),
                        "source_uri": f"FAST_FORMULA_KNOWLEDGE.json#{key}",
                        "title": _clean_title(key),
                        "module": FusionModule.HCM.value,
                        "task_type": task_type,
                        "doc_type": DocType.FAST_FORMULA_DOC,
                        "content": "\n".join(section_lines[:120]),
                        "formula_type": _formula_type_from_text(key, known_formula_types),
                        "use_case": _clean_title(key),
                    },
                    source_file=FORMULA_KNOWLEDGE_PATH.name,
                    source_sheet=key,
                    confidence=0.84,
                )
            )
            notes["classified_assets"]["fast_formula_supporting_docs"].append(
                f"{FORMULA_KNOWLEDGE_PATH}#{key}"
            )

    if FORMULA_CSV_PATH.exists():
        notes["candidate_assets_found"].append(str(FORMULA_CSV_PATH))
        for block in _parse_formula_csv_blocks(FORMULA_CSV_PATH):
            valid, reason = verifier.verify_fast_formula(block["content"])
            if not valid:
                rejected.append({"source_path": str(FORMULA_CSV_PATH), "title": block["title"], "reason": reason or "formula_verifier_failed"})
                notes["classified_assets"]["reject_corpus"].append(f"{FORMULA_CSV_PATH}::{block['title']}")
                continue
            accepted.append(
                _normalize_formula_record(
                    {
                        "source_path": str(FORMULA_CSV_PATH),
                        "source_uri": "Fastformula_KT.csv",
                        "title": _clean_title(block["title"]),
                        "module": FusionModule.HCM.value,
                        "task_type": TaskType.FAST_FORMULA_GENERATION.value,
                        "doc_type": DocType.FAST_FORMULA_EXAMPLE,
                        "content": block["content"],
                        "formula_type": _formula_type_from_text(f"{block['title']} {block['content']}", known_formula_types),
                        "use_case": _clean_title(block["title"]),
                    },
                    source_file=FORMULA_CSV_PATH.name,
                    source_sheet="csv_formula_blocks",
                    confidence=0.89,
                )
            )
            notes["classified_assets"]["fast_formula_examples"].append(f"{FORMULA_CSV_PATH}::{block['title']}")

    for path in FORMULA_SUPPORTING_JSON_PATHS:
        if not path.exists():
            continue
        notes["candidate_assets_found"].append(str(path))
        text = _read_text(path)
        payload = json.loads(text)
        formula_lines = [
            line
            for line in _flatten_section(payload)
            if "fast formula" in line.lower() or "database item" in line.lower() or "compile" in line.lower()
        ]
        if formula_lines:
            accepted.append(
                _normalize_formula_record(
                    {
                        "source_path": str(path),
                        "source_uri": path.name,
                        "title": _clean_title(path.stem),
                        "module": FusionModule.HCM.value,
                        "task_type": TaskType.FAST_FORMULA_TROUBLESHOOTING.value,
                        "doc_type": DocType.FAST_FORMULA_DOC,
                        "content": "\n".join(formula_lines[:120]),
                        "formula_type": _formula_type_from_text(path.stem, known_formula_types),
                        "use_case": _clean_title(path.stem),
                    },
                    source_file=path.name,
                    source_sheet="json_support",
                    confidence=0.8,
                )
            )
            notes["classified_assets"]["fast_formula_supporting_docs"].append(str(path))
        else:
            notes["classified_assets"]["reject_corpus"].append(str(path))

    for path in FORMULA_SUPPORTING_TEXT_PATHS:
        if not path.exists():
            continue
        notes["candidate_assets_found"].append(str(path))
        text = _read_text(path)
        lines = [
            _normalize_formula_line(line)
            for line in text.splitlines()
            if any(token in line.lower() for token in ("fast formula", "formula_", "formula ", "database item", "ff_"))
        ]
        compact_lines = [line for line in lines if line][:120]
        if compact_lines:
            accepted.append(
                _normalize_formula_record(
                    {
                        "source_path": str(path),
                        "source_uri": path.name,
                        "title": _clean_title(path.stem),
                        "module": FusionModule.HCM.value,
                        "task_type": TaskType.FAST_FORMULA_TROUBLESHOOTING.value,
                        "doc_type": DocType.FAST_FORMULA_DOC,
                        "content": "\n".join(compact_lines),
                        "formula_type": _formula_type_from_text(path.stem, known_formula_types),
                        "use_case": _clean_title(path.stem),
                    },
                    source_file=path.name,
                    source_sheet="text_support",
                    confidence=0.78,
                )
            )
            notes["classified_assets"]["fast_formula_supporting_docs"].append(str(path))
        else:
            notes["classified_assets"]["reject_corpus"].append(str(path))

    if FORMULA_TYPES_PATH.exists():
        notes["candidate_assets_found"].append(str(FORMULA_TYPES_PATH))
    for formula_type in known_formula_types:
        accepted.append(
            _normalize_formula_record(
                {
                    "source_path": str(FORMULA_TYPES_PATH),
                    "source_uri": "formula_types.json",
                    "title": f"{formula_type} Formula Type",
                    "module": FusionModule.HCM.value,
                    "task_type": TaskType.FAST_FORMULA_GENERATION.value,
                    "doc_type": DocType.FAST_FORMULA_DOC,
                    "content": f"Oracle Fast Formula type: {formula_type}",
                    "formula_type": formula_type,
                    "use_case": formula_type,
                },
                source_file=FORMULA_TYPES_PATH.name,
                source_sheet="formula_types",
                confidence=0.76,
            )
        )
        notes["classified_assets"]["fast_formula_supporting_docs"].append(f"{FORMULA_TYPES_PATH}::{formula_type}")

    workbook_candidates = _candidate_formula_workbooks()
    notes["candidate_assets_found"].extend(str(path) for path in workbook_candidates)
    notes["excel_formula_example_files_found"] = len(workbook_candidates)
    notes["excel_formula_example_files"] = [str(path) for path in workbook_candidates[:20]]
    workbook_recovery_notes: List[Dict[str, Any]] = []
    for workbook_path in workbook_candidates:
        recovered, workbook_rejects, workbook_notes = _recover_formula_examples_from_workbook(
            workbook_path,
            verifier,
            known_formula_types,
        )
        accepted.extend(recovered)
        rejected.extend(workbook_rejects)
        if recovered:
            notes["classified_assets"]["fast_formula_examples"].append(str(workbook_path))
        else:
            notes["classified_assets"]["reject_corpus"].append(str(workbook_path))
        workbook_recovery_notes.append(workbook_notes)

    notes["workbook_recovery"] = workbook_recovery_notes
    notes["candidate_assets_found"] = sorted(set(notes["candidate_assets_found"]))
    for key, values in notes["classified_assets"].items():
        notes["classified_assets"][key] = sorted(set(values))
    notes["classification_counts"] = {
        key: len(values) for key, values in notes["classified_assets"].items()
    }
    notes["recovered_csv_formula_blocks"] = len(
        [item for item in notes["classified_assets"]["fast_formula_examples"] if str(FORMULA_CSV_PATH) in item]
    )
    notes["true_grounded_formula_examples"] = sum(
        1
        for record in accepted
        if getattr(record.get("doc_type"), "value", record.get("doc_type")) == DocType.FAST_FORMULA_EXAMPLE.value
    )

    return accepted, rejected, notes


def _build_fast_formula_chunks(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    chunks: List[Dict[str, Any]] = []
    seen_hashes = set()

    for record in records:
        content_hash = stable_hash("fast_formula_corpus", record["content"])
        if content_hash in seen_hashes:
            continue
        seen_hashes.add(content_hash)

        content = record["content"]
        formula_name = _extract_formula_name(record["title"], content)
        formula_type = record.get("formula_type") or "UNKNOWN"
        metadata = {
            "formula_name": formula_name,
            "formula_type": formula_type,
            "input_values": record.get("input_values") or _extract_formula_inputs(content),
            "database_items": record.get("database_items") or _extract_formula_database_items(content),
            "contexts": record.get("contexts") or _extract_formula_contexts(content),
            "functions": record.get("functions") or _extract_formula_functions(content),
            "output_pattern": record.get("return_behavior") or _extract_formula_output_pattern(content),
            "use_case": record.get("use_case") or record["title"],
            "default_handling": record.get("default_handling") or _extract_formula_defaults(content),
            "source_file": record.get("source_file") or Path(str(record["source_path"])).name,
            "source_sheet": record.get("source_sheet") or "",
            "derived_from_doc": bool(record.get("derived_from_doc")),
            "source_doc": str(record.get("source_doc") or ""),
            "confidence_band": str(record.get("confidence_band") or ""),
            "authority_tier": "secondary",
        }
        trusted_objects = metadata["database_items"][:20]

        document = CuratedIngestionValidator.build_document(
            source_path=record["source_path"],
            source_uri=record["source_uri"],
            title=record["title"],
            module=record["module"],
            task_type=record["task_type"],
            doc_type=record["doc_type"],
            trusted_schema_objects=trusted_objects,
            quality_score=float(record.get("confidence") or (0.9 if record["doc_type"] == DocType.FAST_FORMULA_EXAMPLE else 0.82)),
            source_system=SourceSystem.ORACLEWINGS_REPO,
            content=content,
            document_id=f"FF::{content_hash[:16]}",
            metadata=metadata,
        )
        chunk = CuratedIngestionValidator.build_chunk(document, content, 0)
        chunks.append(CuratedIngestionValidator.chunk_payload(chunk))

    return chunks


def build_specialization_tracks(tenant_id: str = "demo", reset_indexes: bool = True) -> Dict[str, Any]:
    from core.retrieval.vectors.faiss_index import FaissIndex

    registry = get_default_registry()
    verifier = Verifier()
    SPECIALIZATION_DIR.mkdir(parents=True, exist_ok=True)
    MANIFEST_DIR.mkdir(parents=True, exist_ok=True)

    formula_records, formula_rejects, formula_notes = _inventory_fast_formula_assets(verifier)
    known_formula_types = _load_formula_types()
    harvest_formula_examples = _load_harvest_formula_examples(known_formula_types)
    derived_formula_records, derived_formula_rejects, derived_formula_notes = _derive_formula_examples_from_harvest(
        verifier,
        known_formula_types,
    )
    formula_records = _dedupe_formula_records(
        formula_records + harvest_formula_examples + derived_formula_records
    )

    sql_records, sql_rejects = _inventory_sql_examples(registry, verifier)
    oraclewings_sql_records, oraclewings_sql_rejects = _inventory_oraclewings_sql_examples(registry, verifier)
    derived_sql_records, derived_sql_rejects, derived_sql_notes = _derive_sql_patterns_from_harvest(
        registry,
        verifier,
    )
    join_sql_records, join_sql_rejects, join_sql_notes = _derive_join_graph_sql_patterns(
        registry,
        verifier,
        sql_records + oraclewings_sql_records + derived_sql_records,
    )
    sql_records = _dedupe_sql_records(sql_records + oraclewings_sql_records + derived_sql_records + join_sql_records)

    sql_chunks = _build_sql_example_chunks(sql_records)
    schema_chunks = _build_schema_metadata_chunks(registry, sql_records)
    formula_chunks = _build_fast_formula_chunks(formula_records)
    sql_rejects = sql_rejects + oraclewings_sql_rejects + derived_sql_rejects + join_sql_rejects
    formula_rejects = formula_rejects + derived_formula_rejects

    inventory = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "tenant_id": tenant_id,
        "asset_inventory": {
            "sql_examples_corpus": {
                "accepted_examples": len(sql_records),
                "accepted_source_files": len({record["source_path"] for record in sql_records}),
                "module_breakdown": dict(Counter(record["module"] for record in sql_records)),
                "task_type_breakdown": dict(Counter(record["task_type"] for record in sql_records)),
                "join_rich_patterns": sum(1 for record in sql_records if len(record.get("tables_used", [])) > 1),
            },
            "schema_metadata_corpus": {
                "accepted_chunks": len(schema_chunks),
                "referenced_tables": len({table for record in sql_records for table in record["tables_used"]}),
            },
            "fast_formula_corpus": {
                "accepted_records": len(formula_records),
                "accepted_chunks": len(formula_chunks),
                "doc_type_breakdown": dict(Counter(str(record["doc_type"].value) for record in formula_records)),
                **formula_notes,
                "harvest_real_examples_loaded": len(harvest_formula_examples),
                "derived_formula_examples_added": len(derived_formula_records),
                "derived_formula_notes": derived_formula_notes,
            },
            "harvest_grounding_expansion": {
                "oraclewings_sql_examples_loaded": len(oraclewings_sql_records),
                "derived_sql_patterns_added": len(derived_sql_records) + len(join_sql_records),
                "derived_formula_examples_added": len(derived_formula_records),
                "sql_derivation_notes": {
                    **derived_sql_notes,
                    **join_sql_notes,
                },
                "formula_derivation_notes": derived_formula_notes,
            },
            "reject_corpus": {
                "rejected_sql_records": len(sql_rejects),
                "rejected_fast_formula_records": len(formula_rejects),
                "top_reject_reasons": dict(
                    Counter(item["reason"] for item in sql_rejects + formula_rejects).most_common(20)
                ),
            },
        },
    }

    _write_json(SPECIALIZATION_DIR / "asset_inventory_summary.json", inventory)
    _write_jsonl(MANIFEST_DIR / "sql_examples_manifest.jsonl", sql_chunks)
    _write_jsonl(MANIFEST_DIR / "schema_metadata_manifest.jsonl", schema_chunks)
    _write_jsonl(MANIFEST_DIR / "fast_formula_manifest.jsonl", formula_chunks)
    _write_jsonl(MANIFEST_DIR / "reject_manifest.jsonl", sql_rejects + formula_rejects)

    corpus_payloads = {
        "sql_examples_corpus": sql_chunks,
        "schema_metadata_corpus": schema_chunks,
        "fast_formula_corpus": formula_chunks,
    }
    index_stats = {}
    for corpus_name, chunks in corpus_payloads.items():
        index = FaissIndex(tenant_id=tenant_id, indexes_dir=str(INDEXES_DIR), corpus=corpus_name)
        if reset_indexes:
            index.reset()
        if chunks:
            index.add_chunks_list(chunks, batch_size=32)
        index_stats[corpus_name] = index.stats()

    summary = {
        **inventory,
        "index_stats": index_stats,
        "sql_records": sql_records,
        "formula_records": formula_records,
        "schema_chunk_count": len(schema_chunks),
        "derived_sql_records": derived_sql_records + join_sql_records,
        "derived_formula_records": derived_formula_records,
    }
    _write_json(SPECIALIZATION_DIR / "specialization_ingestion_summary.json", summary)
    return summary
